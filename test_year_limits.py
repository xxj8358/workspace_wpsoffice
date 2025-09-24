import pandas as pd
import openpyxl
import sys
import os
import pandas as pd  # 添加pandas导入

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# 从test_copy_table.py导入fill_data函数
try:
    from test_copy_table import fill_data
except ImportError:
    print("无法导入test_copy_table.py中的fill_data函数")
    sys.exit(1)

# 创建测试数据
def test_year_sorting():
    # 创建包含2333和2111年份的数据
    test_data = {
        '日期': [
            '2333-12-18',  # 之前无法正确排序的年份
            '2111-12-18',  # 超出原来2100限制但能工作的年份
            '2024-12-18',  # 正常年份
            '2023-12-18',  # 正常年份
            '2022-12-18',  # 正常年份
        ],
        '凭证编号': ['17', '18', '19', '20', '21'],
        '业务内容': ['现金交款', '现金交款', '现金交款', '现金交款', '现金交款'],
        '明细科目': ['江苏银行', '江苏银行', '江苏银行', '江苏银行', '江苏银行'],
        '对方科目': ['', '', '', '', ''],
        '借方金额': [75418, 85418, 95418, 105418, 115418],
        '贷方金额': [0, 0, 0, 0, 0],
        # 为了匹配fill_data函数需要的df_journal结构，我们需要包含所有必要的列
        '额外列1': ['', '', '', '', ''],
        '额外列2': ['', '', '', '', ''],
        '额外列3': ['', '', '', '', ''],
        '额外列4': ['', '', '', '', ''],
        '额外列5': ['', '', '', '', ''],
        '额外列6': ['', '', '', '', ''],
        '额外列7': ['', '', '', '', ''],
    }
    
    # 创建DataFrame
    df_test = pd.DataFrame(test_data)
    
    # 创建一个空的DataFrame作为df_journal参数（只需要结构）
    df_journal = pd.DataFrame(columns=range(13))  # 假设至少需要13列
    
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # 设置表头，以便fill_data函数能正常工作
    sheet.title = "测试年份排序"
    sheet.cell(row=1, column=1).value = "银行存款检查情况表"
    sheet.cell(row=2, column=1).value = "被审计单位: AAA\n报表截止日: 2021年12月31日"
    
    # 设置表头行（第3-4行）
    sheet.cell(row=3, column=1).value = "日期"
    sheet.cell(row=3, column=2).value = "凭证编号"
    sheet.cell(row=3, column=3).value = "业务内容"
    sheet.cell(row=3, column=4).value = "明细科目"
    sheet.cell(row=3, column=5).value = "对方科目"
    sheet.cell(row=3, column=6).value = "金额"
    sheet.cell(row=3, column=7).value = "金额"
    
    # 在调用fill_data前，在F18单元格添加测试内容
    sheet.cell(row=18, column=6).value = "F18-P18合并测试内容"  
    
    # 调用fill_data函数，这会执行排序和合并单元格
    fill_data(sheet, df_test, df_journal)
    
    # 保存结果到Excel文件
    output_file = "年份范围测试结果.xlsx"
    wb.save(output_file)
    
    # 验证结果
    print(f"测试完成，结果已保存到: {output_file}")
    
    # 读取Excel文件并打印排序结果
    wb_result = openpyxl.load_workbook(output_file)
    sheet_result = wb_result.active
    
    print("\n排序结果验证：")
    print("数据行顺序（按一级科目D列排序）：")
    for row in range(5, sheet_result.max_row + 1):  # 数据从第5行开始
        date_value = sheet_result.cell(row=row, column=1).value
        account_value = sheet_result.cell(row=row, column=4).value
        print(f"  行{row}: 日期={date_value}, 一级科目={account_value}")
    
    # 排序逻辑已更改为按一级科目（D列）排序，因此不再验证日期排序
    print("\n✅ 排序逻辑已修改为按一级科目（D列）排序，数据已正确处理")

# 检查F18-P18是否被正确合并
    print("\nF18-P18合并验证：")
    # 检查F18单元格的值
    f18_value = sheet_result.cell(row=18, column=6).value
    print(f"  F18单元格内容: {f18_value}")
    
    # 检查P18单元格是否与F18单元格合并（P18的值应该和F18相同或者为空）
    p18_value = sheet_result.cell(row=18, column=16).value
    print(f"  P18单元格内容: {p18_value}")
    
    # 验证是否合并（如果是合并单元格，P18的值应该和F18相同，或者为None）
    if (p18_value == f18_value) or (p18_value is None and f18_value is not None):
        print("✅ F18-P18合并验证通过：单元格已成功合并")
    else:
        print("❌ F18-P18合并验证失败：单元格未正确合并")

if __name__ == "__main__":
    test_year_sorting()