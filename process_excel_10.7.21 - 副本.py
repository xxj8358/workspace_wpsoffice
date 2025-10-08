import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
import os
from copy import copy
import traceback
import tkinter as tk
from tkinter import filedialog
import sys

def process_excel_file(file_path):
    print(f"开始处理文件: {file_path}")
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在")
        return None
    
    try:
        # 读取Excel文件
        xl = pd.ExcelFile(file_path)
        print(f"文件包含工作表: {xl.sheet_names}")
        
        # 读取序时账工作表
        print("\n读取序时账工作表...")
        ledger_df = pd.read_excel(file_path, sheet_name='序时账')
        print(f"序时账数据形状: {ledger_df.shape}")
        print("序时账列名:", list(ledger_df.columns))
        
        # 读取余额表1
        print("\n读取余额表1工作表...")
        balance_df = pd.read_excel(file_path, sheet_name='余额表1')
        print(f"余额表1数据形状: {balance_df.shape}")
        print("余额表1列名:", list(balance_df.columns))
        
        # 使用openpyxl读取序时账以获取背景色信息（不使用data_only=True，以便获取样式）
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ledger_ws = wb['序时账']
        
        # 获取A列（从第2行开始，跳过标题行）背景色为黄色的行索引
        yellow_rows = []
        for row_idx in range(2, ledger_ws.max_row + 1):
            try:
                cell = ledger_ws[f'A{row_idx}']
                # 检查单元格背景色是否为黄色
                fill = cell.fill
                # 黄色的各种表示方式
                if hasattr(fill, 'start_color') and hasattr(fill.start_color, 'index'):
                    color_index = str(fill.start_color.index)
                    # 处理不同格式的黄色表示
                    if 'FFFF00' in color_index or 'FFFF' in color_index and '00' in color_index:
                        # 映射回DataFrame的索引（DataFrame从0开始）
                        yellow_rows.append(row_idx - 2)
            except Exception as e:
                print(f"检查行 {row_idx} 背景色时出错: {e}")
        
        print(f"找到 {len(yellow_rows)} 行背景色为黄色的数据")
        
        # 创建结果DataFrame来存储所有筛选后的数据
        all_filtered_data = []
        
        # 确保余额表1的列名正确
        # 查找F列对应的科目名称列（假设F列是第6列）
        f_column = balance_df.columns[5]  # F列
        print(f"使用列 '{f_column}' 作为余额表1的F列（一级科目名称）")
        
        # 查找G列对应的借方条数列（假设G列是第7列）
        g_column = balance_df.columns[6]  # G列
        print(f"使用列 '{g_column}' 作为余额表1的G列（借方条数）")
        
        # 查找H列对应的贷方条数列（假设H列是第8列）
        h_column = balance_df.columns[7]  # H列
        print(f"使用列 '{h_column}' 作为余额表1的H列（贷方条数）")
        
        # 查找序时账中对应的列
        # 查找F列（一级科目名称，假设F列是第6列）
        ledger_f_column = ledger_df.columns[5]  # F列
        print(f"使用列 '{ledger_f_column}' 作为序时账的F列（一级科目名称）")
        
        # 查找L列（借方发生额，假设L列是第12列）
        ledger_l_column = ledger_df.columns[11]  # L列
        print(f"使用列 '{ledger_l_column}' 作为序时账的L列（借方发生额）")
        
        # 查找M列（贷方发生额，假设M列是第13列）
        ledger_m_column = ledger_df.columns[12]  # M列
        print(f"使用列 '{ledger_m_column}' 作为序时账的M列（贷方发生额）")
        
        # 步骤1: 根据余额表1的内容筛选序时账数据
        print("\n=== 开始筛选数据 ===")
        
        # 创建黄色背景行的标记数组
        is_yellow = [False] * len(ledger_df)
        for idx in yellow_rows:
            if 0 <= idx < len(ledger_df):
                is_yellow[idx] = True
        ledger_df['is_yellow'] = is_yellow
        
        # 遍历余额表1的每一行（不包含第一行标题）
        for idx, row in balance_df.iterrows():
            # 获取一级科目名称
            first_level_name = row[f_column]
            if pd.isna(first_level_name):
                continue
                
            print(f"\n处理一级科目: {first_level_name}")
            
            # 筛选条件1: 该一级科目下的所有数据
            level_data = ledger_df[ledger_df[ledger_f_column] == first_level_name].copy()
            print(f"条件1筛选后: {len(level_data)} 条记录")
            
            # 获取需要筛选的数量
            debit_count = int(row[g_column]) if pd.notna(row[g_column]) else 0
            credit_count = int(row[h_column]) if pd.notna(row[h_column]) else 0
            
            print(f"需要筛选的借方数量: {debit_count}, 贷方数量: {credit_count}")
            
            # 筛选结果存储
            filtered_data = pd.DataFrame()
            
            # 筛选条件2: 按G列数量筛选序时账L列（借方发生额）
            if debit_count > 0:
                # 只筛选有借方发生额的记录
                has_debit = level_data[level_data[ledger_l_column] > 0].copy()
                actual_debit_count = min(debit_count, len(has_debit))
                if actual_debit_count > 0:
                    # 按借方发生额降序排列，取前actual_debit_count条
                    debit_filtered = has_debit.nlargest(actual_debit_count, ledger_l_column)
                    print(f"条件2筛选后，实际筛选借方记录: {len(debit_filtered)} 条")
                    filtered_data = pd.concat([filtered_data, debit_filtered])
                else:
                    print(f"条件2筛选后，没有符合条件的借方记录")
            else:
                print(f"条件2筛选后，debit_count为0，不筛选借方记录")
            
            # 筛选条件3: 按H列数量筛选序时账M列（贷方发生额）
            if credit_count > 0:
                # 只筛选有贷方发生额的记录
                has_credit = level_data[level_data[ledger_m_column] > 0].copy()
                actual_credit_count = min(credit_count, len(has_credit))
                if actual_credit_count > 0:
                    # 按贷方发生额降序排列，取前actual_credit_count条
                    credit_filtered = has_credit.nlargest(actual_credit_count, ledger_m_column)
                    print(f"条件3筛选后，实际筛选贷方记录: {len(credit_filtered)} 条")
                    filtered_data = pd.concat([filtered_data, credit_filtered])
                else:
                    print(f"条件3筛选后，没有符合条件的贷方记录")
            else:
                print(f"条件3筛选后，credit_count为0，不筛选贷方记录")
            
            # 筛选条件4: 添加黄色背景的行
            yellow_data = level_data[level_data['is_yellow'] == True].copy()
            if not yellow_data.empty:
                print(f"条件4筛选后，找到黄色背景记录: {len(yellow_data)} 条")
                # 将黄色背景记录添加到筛选结果中
                filtered_data = pd.concat([filtered_data, yellow_data])
            
            # 去重
            filtered_data = filtered_data.drop_duplicates()
            print(f"该科目最终筛选结果: {len(filtered_data)} 条记录")
            
            # 添加到总结果中
            if not filtered_data.empty:
                all_filtered_data.append(filtered_data)
        
        # 移除临时列
        if 'is_yellow' in ledger_df.columns:
            ledger_df = ledger_df.drop('is_yellow', axis=1)
        
        # 合并所有筛选结果
        if all_filtered_data:
            final_filtered = pd.concat(all_filtered_data)
            print(f"\n总筛选结果: {len(final_filtered)} 条记录")
        else:
            print("\n警告: 没有找到符合条件的记录")
            return None
        
        # 步骤2: 按照一级科目名称拆分到不同工作表
        print("\n=== 开始拆分数据 ===")
        
        # 创建新的Excel文件用于保存结果
        # 定义结果文件路径
        result_file = f"{os.path.splitext(file_path)[0]}_处理结果.xlsx"
        
        # 查找模板工作表
        template_loaded = False
        template_sheet = None
        
        try:
            # 重新加载工作簿以查找模板
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                if '模板' in sheet_name:
                    template_sheet = wb[sheet_name]
                    print(f"找到模板工作表: {template_sheet.title}")
                    
                    template_loaded = True
                    print(f"成功加载模板工作表: {template_sheet.title}")
                    break
            wb.close()
        except Exception as e:
            print(f"加载模板时出错: {e}")
        
        # 创建新工作簿
        result_wb = openpyxl.Workbook()
        # 删除默认的sheet
        default_sheet = result_wb.active
        result_wb.remove(default_sheet)
        
        # 从原始文件复制指定的工作表
        try:
            # 再次打开原始工作簿以复制工作表
            source_wb = openpyxl.load_workbook(file_path, data_only=True)  # 使用data_only=True以获取单元格显示的值
            
            # 打印所有可用的工作表名称，以便调试
            print(f"原始文件中的所有工作表: {source_wb.sheetnames}")
            
            # 先复制特定工作表：基本信息、序时账、余额表1、余额表2，尽可能保留原始格式
            sheets_to_copy = ['基本信息', '序时账', '余额表1', '余额表2']
            
            # 精确匹配复制特定工作表
            for sheet_name in sheets_to_copy:
                if sheet_name in source_wb.sheetnames:
                    # 创建新工作表
                    new_sheet = result_wb.create_sheet(sheet_name)
                    source_sheet = source_wb[sheet_name]
                    
                    # 复制单元格的值和格式
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                            # 复制值
                            new_cell.value = cell.value
                            
                            # 尝试复制格式（如果存在）
                            try:
                                if cell.has_style:
                                    # 复制字体
                                    if cell.font:
                                        new_cell.font = copy(cell.font)
                                    # 复制填充
                                    if cell.fill:
                                        new_cell.fill = copy(cell.fill)
                                    # 复制边框
                                    if cell.border:
                                        new_cell.border = copy(cell.border)
                                    # 复制对齐
                                    if cell.alignment:
                                        new_cell.alignment = copy(cell.alignment)
                                    # 复制数字格式
                                    new_cell.number_format = cell.number_format
                            except Exception as e:
                                # 如果复制格式失败，继续复制下一个单元格
                                pass
                    
                    # 复制列宽
                    for column in source_sheet.column_dimensions:
                        try:
                            new_sheet.column_dimensions[column].width = source_sheet.column_dimensions[column].width
                        except:
                            pass
                    
                    # 复制行高
                    for row in range(1, source_sheet.max_row + 1):
                        try:
                            if source_sheet.row_dimensions[row].height:
                                new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                        except:
                            pass
                    
                    # 复制合并单元格
                    for merged_range in source_sheet.merged_cells.ranges:
                        try:
                            new_sheet.merge_cells(str(merged_range))
                        except:
                            pass
                    
                    print(f"已复制工作表: {sheet_name}")
            
            source_wb.close()
        except Exception as e:
            print(f"复制工作表时出错: {e}")
        
        # 复制筛选后的总数据到一个新工作表
        result_wb.create_sheet('筛选结果')
        result_sheet = result_wb['筛选结果']
        
        # 写入标题行
        for col_idx, col_name in enumerate(final_filtered.columns, 1):
            result_sheet.cell(row=1, column=col_idx, value=col_name)
        
        # 写入数据行
        for row_idx, (_, row) in enumerate(final_filtered.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                result_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 从余额表1的F列获取科目名称的排序顺序
        # 查找F列的列名（可能是'科目名称'或其他名称）
        科目名称列 = None
        for col in balance_df.columns:
            # 检查列内容是否包含科目名称
            if any(balance_df[col].astype(str).str.contains('银行存款|非限定净资产|其他收入|管理费用', na=False)):
                科目名称列 = col
                break
        
        # 如果找到科目名称列，则提取排序顺序
        科目排序顺序 = []
        if 科目名称列:
            # 提取非空的科目名称值
            科目排序顺序 = [str(name).strip() for name in balance_df[科目名称列].dropna() if str(name).strip()]
            print(f"从余额表1提取的科目排序顺序: {科目排序顺序}")
        else:
            print("未找到余额表1中的科目名称列，使用默认排序")
        
        # 按一级科目名称分组
        grouped = final_filtered.groupby(ledger_f_column)
        print(f"共分为 {len(grouped)} 个不同的一级科目")
        
        # 创建按余额表1 F列顺序排序的分组列表
        sorted_groups = []
        # 先添加在余额表1中存在的科目
        for 科目名称 in 科目排序顺序:
            if 科目名称 in grouped.groups:
                sorted_groups.append((科目名称, grouped.get_group(科目名称)))
        # 然后添加余额表1中不存在但在数据中存在的科目
        for name, group_data in grouped:
            if name not in 科目排序顺序:
                sorted_groups.append((name, group_data))
        
        # 计算所有工作表中需要插入的最大行数
        max_rows_to_insert = 0
        original_data_rows = 5  # 原始的5-9行共5行数据区
        
        for name, group_data in sorted_groups:
            data_count = len(group_data)
            rows_to_insert = max(0, data_count - original_data_rows)
            if rows_to_insert > max_rows_to_insert:
                max_rows_to_insert = rows_to_insert
        
        print(f"所有工作表中需要插入的最大行数: {max_rows_to_insert}")
        
        # 按排序后的顺序创建工作表
        for name, group_data in sorted_groups:
            sheet_name = f"{str(name)[:20]}"  # 限制工作表名长度
            
            # 创建新工作表
            new_sheet = result_wb.create_sheet(sheet_name)
            
            # 如果加载了模板，复制模板内容和格式
            if template_loaded and template_sheet:
                # 复制模板工作表的内容和格式
                # 先复制所有行高
                for row_idx in template_sheet.row_dimensions:
                    if row_idx in template_sheet.row_dimensions:
                        new_sheet.row_dimensions[row_idx].height = template_sheet.row_dimensions[row_idx].height
                
                # 复制所有列宽
                for col_idx in template_sheet.column_dimensions:
                    if col_idx in template_sheet.column_dimensions:
                        new_sheet.column_dimensions[col_idx].width = template_sheet.column_dimensions[col_idx].width
                
                # 为P列设置水平靠左对齐
                if 'P' in template_sheet.column_dimensions:
                    for cell in template_sheet['P']:
                        if cell.row >= 3 and cell.row <= 11:  # 仅处理数据区域
                            new_cell = new_sheet[f'P{cell.row}']
                            new_cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 复制单元格的内容、格式、样式等
                for row in template_sheet.iter_rows():
                    for cell in row:
                        # 创建新单元格
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        # 复制单元格样式
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                
                print(f"已应用模板格式和内容到工作表: {sheet_name}")
            
            # 添加指定的单元格合并逻辑
            # 顶部标题行合并
            new_sheet.merge_cells('A1:P1')
            # 设置A1单元格内容为工作表名称加上检查情况表
            new_sheet['A1'] = f'{sheet_name}检查情况表'
            
            # 从模板复制A1单元格的格式
            if template_loaded and template_sheet and 'A1' in template_sheet:
                # 复制模板A1单元格的样式
                if template_sheet['A1'].has_style:
                    new_sheet['A1'].font = copy(template_sheet['A1'].font)
                    new_sheet['A1'].border = copy(template_sheet['A1'].border)
                    new_sheet['A1'].fill = copy(template_sheet['A1'].fill)
                    new_sheet['A1'].number_format = template_sheet['A1'].number_format
                    new_sheet['A1'].protection = copy(template_sheet['A1'].protection)
                    new_sheet['A1'].alignment = copy(template_sheet['A1'].alignment)
            
            # 强制设置A1单元格的字体为宋体18号
            new_sheet['A1'].font = Font(name='宋体', size=18, bold=True)
            new_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 第二行合并（A2到P2）
            new_sheet.merge_cells('A2:P2')
            # 金额列表头合并（F3到G3）
            new_sheet.merge_cells('F3:G3')
            # 核对内容列表头合并（H3到O3）
            new_sheet.merge_cells('H3:O3')
            # 数据列表头合并
            new_sheet.merge_cells('A3:A4')  # 日期列表头
            new_sheet.merge_cells('B3:B4')  # 凭证编号列表头
            new_sheet.merge_cells('C3:C4')  # 业务内容列表头
            new_sheet.merge_cells('D3:D4')  # 明细科目列表头
            new_sheet.merge_cells('E3:E4')  # 对方科目列表头
            new_sheet.merge_cells('P3:P4')  # 备注列表头
           
            # 设置表格外边框为1磅黑实线，保留内部虚线
            # 定义实线样式
            thin_solid_border = Side(style='thin', color='000000')  # 1磅黑色实线
            
            # 获取表格的实际范围（A1到P14）
            start_row, end_row = 1, 14
            start_col, end_col = 1, 16  # A到P
            
            # 遍历整个表格区域，为每个单元格设置边框
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = new_sheet.cell(row=row, column=col)
                    # 创建新的边框对象，初始使用现有边框样式
                    current_border = copy(cell.border) if cell.border else Border()
                    
                    # 确定该单元格需要哪些边框是实线（外边框）
                    top_border = thin_solid_border if row == start_row else current_border.top
                    bottom_border = thin_solid_border if row == end_row else current_border.bottom
                    left_border = thin_solid_border if col == start_col else current_border.left
                    right_border = thin_solid_border if col == end_col else current_border.right
                    
                    # 应用新边框
                    new_sheet.cell(row=row, column=col).border = Border(
                        top=top_border,
                        bottom=bottom_border,
                        left=left_border,
                        right=right_border
                    )
            

   
    
            print(f"\n创建工作表: {sheet_name}, 记录数: {len(group_data)}")
            
            # 对数据先按日期（A列）排序，再按凭证编号（B列）排序
            # 将数据转换为可排序的格式
            sorted_data = []
            for _, row in group_data.iterrows():
                # 获取日期和凭证编号用于排序
                date_value = row.iloc[0] if len(ledger_df.columns) > 0 else ''
                voucher_value = row.iloc[1] if len(ledger_df.columns) > 1 else ''
                sorted_data.append((date_value, voucher_value, row))
            
            # 排序：先按日期，再按凭证编号
            sorted_data.sort(key=lambda x: (x[0], x[1]))
            
            # 计算数据行数
            data_count = len(sorted_data)
            
            # 使用全局计算的最大行数来统一插入
            rows_to_insert = max_rows_to_insert
            
            print(f"在工作表 {sheet_name} 中插入 {rows_to_insert} 行，实际数据量: {data_count}")
            
            # 插入统一数量的行
            if rows_to_insert > 0:
                # 在第9行之后插入额外的行，确保10行及以下的内容下移
                new_sheet.insert_rows(10, amount=rows_to_insert)
            
            # 计算最终的数据区域范围
            start_row = 5
            end_row = start_row + data_count - 1
            
            # 更新合并单元格和边框设置以适应新增的行
            # 移除原有的A11-P11合并并重新设置
            if new_sheet.merged_cells:
                merged_ranges_to_remove = []
                # 首先收集需要移除的合并范围
                for merged_range in new_sheet.merged_cells.ranges:
                    if 'A11:P11' in str(merged_range):
                        merged_ranges_to_remove.append(merged_range)
                
                # 然后移除这些合并范围，使用try-except确保不会出错
                for merged_range in merged_ranges_to_remove:
                    try:
                        new_sheet.unmerge_cells(str(merged_range))
                    except Exception as e:
                        print(f"取消合并单元格时出错: {e}，继续处理")
            
            # 设置新的A11+rows_to_insert:P11+rows_to_insert合并
            try:
                new_sheet.merge_cells(f'A{11+rows_to_insert}:P{11+rows_to_insert}')
            except Exception as e:
                print(f"合并单元格时出错: {e}，继续处理")
            
            # 更新表格外边框
            # 定义实线样式
            thin_solid_border = Side(style='thin', color='000000')  # 1磅黑色实线
            
            # 更新表格范围（A1到P14+rows_to_insert）
            start_table_row = 1
            # 表格结束行需要根据插入的行数进行调整，确保包含所有内容
            end_table_row = 14 + rows_to_insert
            start_col, end_col = 1, 16  # A到P
            
            # 合并表格倒数第二行的A-P列单元格
            last_second_row = end_table_row - 1
            new_sheet.merge_cells(f'A{last_second_row}:P{last_second_row}')
            
            # 合并表格倒数第三行的A-E列单元格
            last_third_row = end_table_row - 2
            new_sheet.merge_cells(f'A{last_third_row}:E{last_third_row}')
            # 合并表格倒数第三行的F-P列单元格
            new_sheet.merge_cells(f'F{last_third_row}:P{last_third_row}')
            # 设置表格倒数第三行的行高为50
            new_sheet.row_dimensions[last_third_row].height = 50
            
            # 复制原始格式到新插入的行
            if rows_to_insert > 0:
                # 复制第9行的格式到所有新插入的行
                for row_idx in range(1, rows_to_insert + 1):
                    # 源行是第9行
                    source_row = 9
                    # 目标行是第9+row_idx行（因为是在第10行插入）
                    target_row = 9 + row_idx
                    
                    for col in range(start_col, end_col + 1):
                        try:
                            # 复制格式
                            source_cell = new_sheet.cell(row=source_row, column=col)
                            target_cell = new_sheet.cell(row=target_row, column=col)
                            
                            # 复制样式
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)
                            
                            # 清空内容，因为内容会在后续步骤中填充
                            target_cell.value = ''
                        except Exception as e:
                            print(f"复制行格式时出错: {e}，继续处理")
            
            # 遍历整个表格区域，为每个单元格设置边框
            for row in range(start_table_row, end_table_row + 1):
                for col in range(start_col, end_col + 1):
                    try:
                        cell = new_sheet.cell(row=row, column=col)
                        # 创建新的边框对象，初始使用现有边框样式
                        current_border = copy(cell.border) if cell.border else Border()
                        
                        # 确定该单元格需要哪些边框是实线
                        top_border = thin_solid_border if row == start_table_row else current_border.top
                        bottom_border = thin_solid_border if row == end_table_row else current_border.bottom
                        left_border = thin_solid_border if col == start_col else current_border.left
                        right_border = current_border.right  # 始终使用当前边框的右侧样式，不强制设置为实线
                        
                        # 应用新边框
                        new_sheet.cell(row=row, column=col).border = Border(
                            top=top_border,
                            bottom=bottom_border,
                            left=left_border,
                            right=right_border
                        )
                    except Exception as e:
                        print(f"设置边框时出错: {e}，继续处理")
            
            # 确保数据区域的内部边框也正确设置
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    try:
                        cell = new_sheet.cell(row=row, column=col)
                        # 创建新的边框对象，使用现有的边框样式
                        current_border = copy(cell.border) if cell.border else Border()
                        
                        # 设置内部单元格的下边框
                        bottom_border = thin_solid_border if row < end_row else current_border.bottom
                        
                        # 应用更新的边框
                        new_sheet.cell(row=row, column=col).border = Border(
                            top=current_border.top,
                            bottom=bottom_border,
                            left=current_border.left,
                            right=current_border.right
                        )
                    except Exception as e:
                        print(f"设置数据区域边框时出错: {e}，继续处理")
            
            # 为第12行设置自动换行功能并允许Excel自动调整行高
            # 直接设置第12行，不考虑插入行数的影响
            target_row = 12
            
            try:
                # 设置第12行A到P列的单元格为自动换行，确保文字过长时能自适应调整行高
                for col in range(1, 17):  # A到P列
                    try:
                        cell = new_sheet.cell(row=target_row, column=col)
                        # 复制当前对齐方式，确保不丢失原有设置
                        current_alignment = copy(cell.alignment) if cell.alignment else Alignment()
                        # 设置自动换行
                        current_alignment.wrap_text = True
                        cell.alignment = current_alignment
                    except Exception as e:
                        print(f"设置单元格自动换行时出错 (行 {target_row}, 列 {col}): {e}，继续处理")
                
                # 移除固定行高设置，让Excel根据内容自动调整行高
                if target_row in new_sheet.row_dimensions:
                    new_sheet.row_dimensions[target_row].height = None
                
                print(f"已为第{target_row}行的A到P列单元格启用自动换行功能，行高将根据内容自动调整")
            except Exception as e:
                print(f"设置第{target_row}行行高时出错: {e}，继续处理")
            
            # 在数据区域（5-9行，如有必要扩展）填写已排序的数据
            for data_idx, (_, _, row) in enumerate(sorted_data, start_row):
                try:
                    # A列：日期（序时账A列，第1列）- 只保留年月日
                    if len(ledger_df.columns) > 0:
                        date_value = row.iloc[0]
                        if pd.notna(date_value):
                            # 尝试将日期格式化为只包含年月日
                            try:
                                if isinstance(date_value, str):
                                    # 如果是字符串，尝试解析为日期
                                    date_obj = pd.to_datetime(date_value)
                                    new_sheet[f'A{data_idx}'] = date_obj.strftime('%Y-%m-%d')
                                else:
                                    # 如果是日期类型，直接格式化
                                    new_sheet[f'A{data_idx}'] = date_value.strftime('%Y-%m-%d')
                            except:
                                # 如果格式化失败，保留原始值
                                new_sheet[f'A{data_idx}'] = str(date_value).split()[0]  # 尝试简单分割去除时分秒
                        else:
                            new_sheet[f'A{data_idx}'] = ''
                    
                    # B列：凭证编号（序时账B列，第2列）
                    if len(ledger_df.columns) > 1:
                        new_sheet[f'B{data_idx}'] = row.iloc[1] if pd.notna(row.iloc[1]) else ''
                    
                    # C列：业务内容（序时账J列，第10列）
                    if len(ledger_df.columns) > 9:
                        new_sheet[f'C{data_idx}'] = row.iloc[9] if pd.notna(row.iloc[9]) else ''
                    
                    # D列：明细科目（序时账G列，第7列）
                    if len(ledger_df.columns) > 6:
                        new_sheet[f'D{data_idx}'] = row.iloc[6] if pd.notna(row.iloc[6]) else ''
                    
                    # E列：对方科目（序时账I列，第9列）
                    if len(ledger_df.columns) > 8:
                        new_sheet[f'E{data_idx}'] = row.iloc[8] if pd.notna(row.iloc[8]) else ''
                    
                    # 计算借贷方金额
                    debit_amount = row.iloc[11] if (len(ledger_df.columns) > 11 and pd.notna(row.iloc[11])) else 0
                    credit_amount = row.iloc[12] if (len(ledger_df.columns) > 12 and pd.notna(row.iloc[12])) else 0
                    # 获取序时账M列数据（第13列，贷方发生额）
                    ledger_m_column_data = row.iloc[12] if (len(ledger_df.columns) > 12 and pd.notna(row.iloc[12])) else ''
                    
                    # F列：借方（序时账L列，第12列）
                    if len(ledger_df.columns) > 11:
                        new_sheet[f'F{data_idx}'] = debit_amount
                    else:
                        new_sheet[f'F{data_idx}'] = ''
                    
                    # G列：对应序时账M列数据（第13列，贷方发生额）
                    new_sheet[f'G{data_idx}'] = ledger_m_column_data
                    
                    # H列：留空
                    new_sheet[f'H{data_idx}'] = ''
                    
                    # I-P列：留空，等待手动填写
                    for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O']:
                        new_sheet[f'{col}{data_idx}'] = ''
                    
                    # 从序时账N列（第14列）设置P列内容
                    if len(ledger_df.columns) > 13:  # 确保有足够的列
                        new_sheet[f'P{data_idx}'] = row.iloc[13] if pd.notna(row.iloc[13]) else ''
                        # 设置P列左对齐
                        new_sheet[f'P{data_idx}'].alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        new_sheet[f'P{data_idx}'] = ''
                        # 设置P列左对齐
                        new_sheet[f'P{data_idx}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 检查并处理合并单元格，确保内容正确填充
                    try:
                        # 在填充数据前，先检查并处理可能的合并单元格
                        # 复制合并单元格列表，避免在迭代过程中修改集合导致错误
                        merged_ranges_list = list(new_sheet.merged_cells.ranges)
                        
                        # 先取消当前数据行的所有合并单元格
                        for merged_range in merged_ranges_list:
                            try:
                                # 获取合并区域的范围
                                min_col, min_row, max_col, max_row = merged_range.bounds
                                
                                # 如果当前数据行在合并区域内
                                if min_row <= data_idx <= max_row:
                                    try:
                                        # 取消合并单元格
                                        new_sheet.unmerge_cells(range_string=str(merged_range))
                                        print(f"已取消行 {data_idx} 的合并单元格: {merged_range}")
                                    except Exception as unmerge_e:
                                        print(f"取消合并单元格时出错: {unmerge_e}")
                            except Exception as range_e:
                                print(f"处理合并区域时出错: {range_e}")
                        
                        # 数据完整性检查和修复：确保关键字段的数据被正确填充
                        # 检查A列（日期）是否为空，如果为空且前一行有值，尝试从数据源重新获取
                        if new_sheet[f'A{data_idx}'].value in [None, '']:
                            if len(ledger_df.columns) > 0 and pd.notna(row.iloc[0]):
                                # 重新填充日期
                                try:
                                    date_value = row.iloc[0]
                                    if isinstance(date_value, str):
                                        date_obj = pd.to_datetime(date_value)
                                        new_sheet[f'A{data_idx}'] = date_obj.strftime('%Y-%m-%d')
                                    else:
                                        new_sheet[f'A{data_idx}'] = date_value.strftime('%Y-%m-%d')
                                    print(f"已修复行 {data_idx} 的日期数据")
                                except:
                                    new_sheet[f'A{data_idx}'] = str(row.iloc[0]).split()[0] if pd.notna(row.iloc[0]) else ''
                        
                        # 检查B列（凭证编号）是否为空
                        if new_sheet[f'B{data_idx}'].value in [None, '']:
                            if len(ledger_df.columns) > 1 and pd.notna(row.iloc[1]):
                                new_sheet[f'B{data_idx}'] = row.iloc[1]
                                print(f"已修复行 {data_idx} 的凭证编号")
                        
                        # 检查C列（业务内容）是否为空
                        if new_sheet[f'C{data_idx}'].value in [None, '']:
                            if len(ledger_df.columns) > 9 and pd.notna(row.iloc[9]):
                                new_sheet[f'C{data_idx}'] = row.iloc[9]
                                print(f"已修复行 {data_idx} 的业务内容")
                        
                        # 检查F列和G列（借贷方金额）是否为空
                        if new_sheet[f'F{data_idx}'].value in [None, ''] and len(ledger_df.columns) > 11:
                            new_sheet[f'F{data_idx}'] = row.iloc[11] if pd.notna(row.iloc[11]) else 0
                            print(f"已修复行 {data_idx} 的借方金额")
                        
                        if new_sheet[f'G{data_idx}'].value in [None, ''] and len(ledger_df.columns) > 12:
                            new_sheet[f'G{data_idx}'] = row.iloc[12] if pd.notna(row.iloc[12]) else 0
                            print(f"已修复行 {data_idx} 的贷方金额")
                        
                        # 额外的数据完整性检查：检查前后行数据是否连贯，如果发现问题则尝试修复
                        # 检查是否存在整行几乎为空但应该有数据的情况
                        if (new_sheet[f'A{data_idx}'].value in [None, ''] and 
                            new_sheet[f'B{data_idx}'].value in [None, ''] and 
                            new_sheet[f'C{data_idx}'].value in [None, ''] and 
                            (new_sheet[f'F{data_idx}'].value not in [None, '', 0] or 
                             new_sheet[f'G{data_idx}'].value not in [None, '', 0])):
                            # 如果金额列有值但关键信息列没有值，尝试重新填充整行
                            print(f"检测到行 {data_idx} 数据不完整，尝试重新填充")
                            
                            # 重新填充关键字段
                            if len(ledger_df.columns) > 0 and pd.notna(row.iloc[0]):
                                try:
                                    date_value = row.iloc[0]
                                    if isinstance(date_value, str):
                                        date_obj = pd.to_datetime(date_value)
                                        new_sheet[f'A{data_idx}'] = date_obj.strftime('%Y-%m-%d')
                                    else:
                                        new_sheet[f'A{data_idx}'] = date_value.strftime('%Y-%m-%d')
                                except:
                                    new_sheet[f'A{data_idx}'] = str(row.iloc[0]).split()[0] if pd.notna(row.iloc[0]) else ''
                            
                            if len(ledger_df.columns) > 1 and pd.notna(row.iloc[1]):
                                new_sheet[f'B{data_idx}'] = row.iloc[1]
                            
                            if len(ledger_df.columns) > 9 and pd.notna(row.iloc[9]):
                                new_sheet[f'C{data_idx}'] = row.iloc[9]
                    except Exception as e:
                        print(f"处理合并单元格和数据完整性时出错: {e}")
                except Exception as e:
                    print(f"填充数据时出错: {e}")
                    traceback.print_exc()
        
        # 定义临时结果文件路径
        temp_result_file = f"{os.path.splitext(file_path)[0]}_处理结果_temp.xlsx"
        result_file = f"{os.path.splitext(file_path)[0]}_处理结果.xlsx"
        
        # 保存结果到临时文件
        result_wb.save(temp_result_file)
        
        # 如果目标文件已存在，尝试删除它
        if os.path.exists(result_file):
            try:
                os.remove(result_file)
            except Exception as e:
                print(f"尝试删除已存在的结果文件时出错: {e}")
        
        # 将临时文件重命名为最终文件名
        try:
            os.rename(temp_result_file, result_file)
            print(f"\n处理完成! 结果已保存至: {result_file}")
            return result_file
        except Exception as e:
            print(f"重命名临时文件时出错: {e}")
            print(f"临时文件已保存至: {temp_result_file}")
            return temp_result_file
        
    except Exception as e:
        print(f"处理文件时出错: {e}")
        traceback.print_exc()
        return None

def select_file():
    """打开文件选择对话框让用户选择Excel文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 设置文件类型过滤
    file_types = [
        ("Excel文件", "*.xlsx;*.xls"),
        ("所有文件", "*.*")
    ]
    
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=file_types
    )
    
    return file_path

if __name__ == "__main__":
    try:
        # 设置tkinter中文显示
        if hasattr(sys, 'frozen'):
            # 对于打包后的可执行文件，可能需要额外设置
            pass
        
        # 直接打开文件选择窗口让用户选择文件
        print("请选择要处理的Excel文件...")
        target_file = select_file()
        
        # 检查用户是否选择了文件
        if not target_file:
            print("未选择文件，程序退出。")
        else:
            print(f"已选择文件: {target_file}")
            result = process_excel_file(target_file)
            if result:
                print(f"\n处理完成，请查看生成的文件: {result}")
            else:
                print("\n处理失败，请查看错误信息。")
    except Exception as e:
        print(f"程序运行出错: {e}")
        traceback.print_exc()
    
    # 等待用户按回车键后退出
    input("\n按回车键退出...")
    
    