import pandas as pd
import os
import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import copy
from datetime import datetime

# 测试复制表格功能
def test_copy_table():
    # 让用户选择要处理的Excel文件
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 打开文件选择对话框
    test_file = filedialog.askopenfilename(
        title="选择要处理的Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    
    if not test_file:
        print("用户取消了文件选择，测试终止")
        return
    
    # 创建输出文件路径
    base_name = os.path.basename(test_file)
    name_without_ext = os.path.splitext(base_name)[0]
    output_dir = os.path.dirname(test_file)
    output_file = os.path.join(output_dir, f"{name_without_ext}_处理结果.xlsx")
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # 检查文件扩展名
        file_ext = os.path.splitext(test_file)[1].lower()
        
        # 根据文件类型选择不同的处理方式
        if file_ext == '.xls':
            # 处理.xls文件，需要使用xlrd库
            import xlrd
            
            # 创建新的工作簿作为输出
            wb_output = openpyxl.Workbook()
            wb_output.remove(wb_output.active)  # 删除默认工作表
            
            # 使用xlrd读取.xls文件
            book = xlrd.open_workbook(test_file)
            
            # 复制源文件的所有工作表到输出文件
            for sheet_name in book.sheet_names():
                # 创建新工作表
                ws = wb_output.create_sheet(sheet_name)
                
                # 获取xlrd工作表
                xlrd_sheet = book.sheet_by_name(sheet_name)
                
                # 复制数据
                for row_idx in range(xlrd_sheet.nrows):
                    for col_idx in range(xlrd_sheet.ncols):
                        # 获取单元格值
                        cell_value = xlrd_sheet.cell_value(row_idx, col_idx)
                        
                        # 处理日期类型
                        if xlrd_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            try:
                                # 转换为datetime对象
                                from datetime import datetime
                                cell_value = datetime(*xlrd.xldate_as_tuple(cell_value, book.datemode))
                            except:
                                # 如果转换失败，保留原始值
                                pass
                        
                        # 写入到openpyxl工作表
                        ws.cell(row=row_idx+1, column=col_idx+1).value = cell_value
            
            # 读取序时账数据（不包含第一行）
            if "序时账" in book.sheet_names():
                xlrd_journal = book.sheet_by_name("序时账")
                # 创建DataFrame
                data = []
                for row_idx in range(1, xlrd_journal.nrows):  # 跳过第一行
                    row_data = []
                    for col_idx in range(xlrd_journal.ncols):
                        cell_value = xlrd_journal.cell_value(row_idx, col_idx)
                        # 处理日期类型
                        if xlrd_journal.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            try:
                                from datetime import datetime
                                cell_value = datetime(*xlrd.xldate_as_tuple(cell_value, book.datemode))
                            except:
                                pass
                        row_data.append(cell_value)
                    data.append(row_data)
                # 获取表头
                headers = [xlrd_journal.cell_value(0, col_idx) for col_idx in range(xlrd_journal.ncols)]
                df_journal = pd.DataFrame(data, columns=headers)
            else:
                print("文件中未找到'序时账'工作表")
                return
        else:
            # 处理.xlsx文件
            # 直接加载原始工作簿
            wb_source = openpyxl.load_workbook(test_file)
            
            # 创建新的工作簿作为输出
            wb_output = openpyxl.Workbook()
            wb_output.remove(wb_output.active)  # 删除默认工作表
            
            # 复制源文件的所有工作表到输出文件
            for sheet_name in wb_source.sheetnames:
                # 创建新工作表
                ws_output = wb_output.create_sheet(sheet_name)
                
                # 复制源工作表内容
                source_sheet = wb_source[sheet_name]
                for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row,
                                                 min_col=1, max_col=source_sheet.max_column):
                    for cell in row:
                        new_cell = ws_output[cell.coordinate]
                        if cell.value is not None:
                            new_cell.value = cell.value
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy.copy(cell.alignment)
            
            # 读取序时账数据（使用pandas）
            xl = pd.ExcelFile(test_file)
            if "序时账" not in xl.sheet_names:
                print("文件中未找到'序时账'工作表")
                return
            df_journal = xl.parse("序时账", skiprows=1)  # 跳过第一行
        
        print(f"读取到序时账数据: {len(df_journal)} 行")
        
        # 复制原始表格到输出文件作为单独的工作表
        copy_original_table(test_file, wb_output)
        
        # 分析、筛选和校验序时账数据
        process_journal_data(df_journal, wb_output)
        
        # 保存处理后的文件
        try:
            wb_output.save(output_file)
            print(f"文件处理成功！输出文件: {output_file}")
        except PermissionError:
            print(f"测试失败: [Errno 13] Permission denied: '{output_file}'。请确保该文件没有被其他程序打开，并具有写入权限。")
            return
        
        # 验证输出文件
        if os.path.exists(output_file):
            print(f"验证成功：输出文件 {output_file} 已创建")
            # 可以添加更多的验证逻辑
        else:
            print(f"测试失败：输出文件 {output_file} 不存在")
            
    except Exception as e:
        print(f"测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭tkinter窗口
        root.destroy()

def copy_original_table(input_file, workbook):
    """复制原始Excel文件中的模板工作表到指定工作簿，完整保留格式"""
    file_ext = os.path.splitext(input_file)[1].lower()
    
    # 添加一个新工作表到主工作簿，默认为"原始表格副本"
    sheet_name = "原始表格副本"
    # 确保工作表名称不重复
    counter = 1
    temp_name = sheet_name
    while temp_name in workbook.sheetnames:
        temp_name = f"{sheet_name}_{counter}"
        counter += 1
    sheet_name = temp_name
    
    # 创建新工作表
    new_sheet = workbook.create_sheet(sheet_name)
    
    # 标识是否已找到模板工作表
    template_found = False
    
    if file_ext == '.xls':
        # 处理.xls文件
        try:
            import xlrd
            book = xlrd.open_workbook(input_file)
            if not book.sheets():
                print("文件中没有工作表")
                return
            
            # 尝试找到模板工作表
            source_sheet = None
            template_keywords = ['模板', 'TEMPLATE', 'temp', 'Temp']
            
            # 首先查找包含模板关键词的工作表
            for sheet in book.sheets():
                sheet_name_lower = sheet.name.lower()
                if any(keyword.lower() in sheet_name_lower for keyword in template_keywords):
                    source_sheet = sheet
                    template_found = True
                    print(f"找到模板工作表: {sheet.name}")
                    break
            
            # 如果没找到模板工作表，使用第一个工作表
            if source_sheet is None:
                source_sheet = book.sheets()[0]  # 使用第一个工作表作为模板
                print(f"未找到明确的模板工作表，使用第一个工作表: {source_sheet.name}")
            
            # 复制单元格内容
            for row_idx in range(source_sheet.nrows):
                for col_idx in range(source_sheet.ncols):
                    # 获取单元格值
                    cell_value = source_sheet.cell_value(row_idx, col_idx)
                    
                    # 处理日期类型
                    if source_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                        try:
                            from datetime import datetime
                            cell_value = datetime(*xlrd.xldate_as_tuple(cell_value, book.datemode))
                        except:
                            # 如果转换失败，保留原始值
                            pass
                    
                    # 写入单元格
                    new_cell = new_sheet.cell(row=row_idx+1, column=col_idx+1)
                    if cell_value is not None:
                        new_cell.value = cell_value
            
            # 对于.xls文件，尝试通过转换为.xlsx后再复制格式（如果可能）
            try:
                import tempfile
                import shutil
                from openpyxl import load_workbook
                
                # 创建临时.xlsx文件
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_xlsx:
                    temp_xlsx_path = temp_xlsx.name
                
                # 使用pandas将.xls转换为.xlsx以保留更多格式
                # 尝试只转换我们选定的模板工作表
                df = pd.read_excel(input_file, sheet_name=source_sheet.name)
                with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='模板')
                
                # 加载转换后的.xlsx文件
                temp_wb = load_workbook(temp_xlsx_path)
                if temp_wb.worksheets:
                    temp_sheet = temp_wb['模板']
                    
                    # 复制转换后的格式
                    for row in temp_sheet.iter_rows(min_row=1, max_row=min(temp_sheet.max_row, source_sheet.nrows)):
                        for cell in row:
                            if cell.coordinate in new_sheet and cell.has_style:
                                new_cell = new_sheet[cell.coordinate]
                                new_cell.font = copy.copy(cell.font)
                                new_cell.border = copy.copy(cell.border)
                                new_cell.fill = copy.copy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.alignment = copy.copy(cell.alignment)
                
                # 清理临时文件
                try:
                    os.remove(temp_xlsx_path)
                except:
                    pass
                
                print(f"成功复制模板工作表内容和部分格式到工作表 '{sheet_name}'")
            except Exception as inner_e:
                print(f"尝试复制模板工作表格式时出错: {str(inner_e)}")
                print(f"成功复制模板工作表内容到工作表 '{sheet_name}'")
        except Exception as e:
            print(f"复制模板工作表时出错: {str(e)}")
    else:
        # 处理.xlsx文件
        try:
            # 加载原始文件
            original_wb = openpyxl.load_workbook(input_file)
            
            # 尝试找到模板工作表
            source_sheet = None
            template_keywords = ['模板', 'TEMPLATE', 'temp', 'Temp']
            
            # 首先查找包含模板关键词的工作表
            for sheet_name_xlsx in original_wb.sheetnames:
                if any(keyword.lower() in sheet_name_xlsx.lower() for keyword in template_keywords):
                    source_sheet = original_wb[sheet_name_xlsx]
                    template_found = True
                    print(f"找到模板工作表: {sheet_name_xlsx}")
                    break
            
            # 如果没找到模板工作表，使用第一个工作表
            if source_sheet is None:
                source_sheet = original_wb.worksheets[0]  # 使用第一个工作表作为模板
                print(f"未找到明确的模板工作表，使用第一个工作表: {source_sheet.title}")
            
            # 复制单元格内容和格式
            for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, 
                                             min_col=1, max_col=source_sheet.max_column):
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    if cell.value is not None:
                        new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy.copy(cell.font)
                        new_cell.border = copy.copy(cell.border)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy.copy(cell.protection)
                        new_cell.alignment = copy.copy(cell.alignment)
                
            # 复制列宽设置
            for col in source_sheet.column_dimensions:
                if col in new_sheet.column_dimensions:
                    new_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
                else:
                    new_dim = copy.copy(source_sheet.column_dimensions[col])
                    new_dim.parent = new_sheet
                    new_sheet.column_dimensions[col] = new_dim
                
            # 复制行高设置
            for row in source_sheet.row_dimensions:
                if row in new_sheet.row_dimensions:
                    new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                else:
                    new_dim = copy.copy(source_sheet.row_dimensions[row])
                    new_dim.parent = new_sheet
                    new_sheet.row_dimensions[row] = new_dim
                
            # 复制合并单元格
            for merged_range in source_sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))
                
            # 如果找到的是真正的模板工作表，添加标记
            if template_found:
                print(f"成功复制模板工作表 '{source_sheet.title}' 到工作表 '{sheet_name}'")
            else:
                print(f"成功复制第一个工作表作为模板到工作表 '{sheet_name}'")
        except Exception as e:
            print(f"复制模板工作表时出错: {str(e)}")

def process_journal_data(df_journal, workbook):
    """处理序时账数据，创建科目代码工作表，保留原始模板格式"""
    # 打印列信息以便调试
    print("\n--- 序时账列信息 ---\n")
    for i, col in enumerate(df_journal.columns):
        print(f"列{i+1} ({chr(65+i)}列): '{col}' - 示例值: {df_journal[col].iloc[0] if len(df_journal) > 0 else '空'}")
    
    # 使用第6列（F列）作为科目编码列（索引为5）
    if len(df_journal.columns) >= 6:
        code_column = df_journal.columns[5]  # 第6列（索引为5）
        print(f"\n使用第6列(F列) '{code_column}' 作为科目编码列")
    else:
        print("序时账列数不足，请检查文件格式")
        return
    
    # 使用第5列（E列）作为借/贷列，第12列（L列）作为借方金额列，第13列（M列）作为贷方金额列
    if len(df_journal.columns) >= 13:
        debit_credit_column = df_journal.columns[4]  # 第5列（E列）
        debit_column = df_journal.columns[11]        # 第12列（L列）
        credit_column = df_journal.columns[12]       # 第13列（M列）
        print(f"使用第5列(E列) '{debit_credit_column}' 作为借/贷列")
        print(f"使用第12列(L列) '{debit_column}' 作为借方金额列")
        print(f"使用第13列(M列) '{credit_column}' 作为贷方金额列")
    else:
        print("序时账列数不足13列，请检查文件格式")
        return
    
    # 尝试将金额列转换为数值类型 - 增强版处理
    try:
        # 先检查是否存在'借方发生额'列，如果存在，使用它替代默认的金额列
        if '借方发生额' in df_journal.columns:
            print("检测到'借方发生额'列，使用它替代默认的金额列")
            debit_column = '借方发生额'
            if '贷方发生额' in df_journal.columns:
                credit_column = '贷方发生额'
                print("同时检测到'贷方发生额'列")
        
        # 先尝试直接转换
        df_journal[debit_column] = pd.to_numeric(df_journal[debit_column], errors='coerce')
        df_journal[credit_column] = pd.to_numeric(df_journal[credit_column], errors='coerce')
        
        # 检查转换结果
        if df_journal[debit_column].isna().all() or df_journal[credit_column].isna().all():
            print("直接转换失败，尝试特殊字符处理...")
            # 如果全部转为NaN，尝试处理可能的特殊格式（如千分位，¥符号等）
            if df_journal[debit_column].dtype == 'object':
                df_journal[debit_column] = df_journal[debit_column].astype(str).str.replace(',', '').str.replace('¥', '').str.replace(' ', '')
                df_journal[debit_column] = pd.to_numeric(df_journal[debit_column], errors='coerce')
            
            if df_journal[credit_column].dtype == 'object':
                df_journal[credit_column] = df_journal[credit_column].astype(str).str.replace(',', '').str.replace('¥', '').str.replace(' ', '')
                df_journal[credit_column] = pd.to_numeric(df_journal[credit_column], errors='coerce')
        
        # 填充NaN值为0
        df_journal[debit_column] = df_journal[debit_column].fillna(0)
        df_journal[credit_column] = df_journal[credit_column].fillna(0)
        
        # 确保最终是数值类型
        df_journal[debit_column] = df_journal[debit_column].astype(float)
        df_journal[credit_column] = df_journal[credit_column].astype(float)
        
        print(f"成功将 '{debit_column}' 和 '{credit_column}' 列转换为数值类型")
        print(f"- 借方列类型: {df_journal[debit_column].dtype}")
        print(f"- 贷方列类型: {df_journal[credit_column].dtype}")
        print(f"- 借方列非零值数量: {(df_journal[debit_column] != 0).sum()}")
        print(f"- 贷方列非零值数量: {(df_journal[credit_column] != 0).sum()}")
        print(f"- 借方列示例值: {df_journal[debit_column].iloc[:5].tolist()}")
        print(f"- 贷方列示例值: {df_journal[credit_column].iloc[:5].tolist()}")
    except Exception as e:
        print(f"转换金额列类型时出错: {str(e)}")
        # 如果转换失败，创建新的数值列作为备选
        df_journal['_temp_debit'] = 0
        df_journal['_temp_credit'] = 0
        
        try:
            # 尝试手动解析
            for i, row in df_journal.iterrows():
                try:
                    # 处理借方金额
                    if isinstance(row[debit_column], str):
                        clean_val = row[debit_column].replace(',', '').replace('¥', '').replace(' ', '')
                        df_journal.at[i, '_temp_debit'] = float(clean_val)
                    elif pd.notna(row[debit_column]):
                        df_journal.at[i, '_temp_debit'] = float(row[debit_column])
                    
                    # 处理贷方金额
                    if isinstance(row[credit_column], str):
                        clean_val = row[credit_column].replace(',', '').replace('¥', '').replace(' ', '')
                        df_journal.at[i, '_temp_credit'] = float(clean_val)
                    elif pd.notna(row[credit_column]):
                        df_journal.at[i, '_temp_credit'] = float(row[credit_column])
                except:
                    pass
            
            # 使用临时列
            debit_column = '_temp_debit'
            credit_column = '_temp_credit'
            print("已创建临时数值列用于排序")
        except:
            print("创建临时列也失败，使用索引排序作为最后的备选方案")
            # 添加一个排序键列
            df_journal['_sort_index'] = range(len(df_journal))
    
    # 获取所有不重复的科目代码
    unique_codes = df_journal[code_column].unique()
    print(f"找到 {len(unique_codes)} 个不同的科目编码")
    
    # 找到模板工作表
    original_sheet = None
    for s in workbook.sheetnames:
        if "模板" in s.lower():
            original_sheet = workbook[s]
            print(f"找到模板工作表: {s}")
            break
    
    # 如果没找到模板工作表，再查找原始表格
    if original_sheet is None:
        for s in workbook.sheetnames:
            if "原始表格" in s:
                original_sheet = workbook[s]
                print(f"找到原始表格工作表: {s}")
                break
    
    # 如果还是没找到，尝试直接使用第一个工作表
    if original_sheet is None and workbook.sheetnames:
        original_sheet = workbook[workbook.sheetnames[0]]
        print(f"未找到明确的模板工作表，使用第一个工作表: {original_sheet.title}")
    
    # 处理每个唯一科目编码
    for code in unique_codes:
        if pd.isna(code):
            continue
            
        code_str = str(code).strip()
        if not code_str:
            continue
            
        print(f"处理科目编码: {code_str}")
        
        # 筛选当前科目编码的数据
        sub_df = df_journal[df_journal[code_column] == code].copy()
        
        if sub_df.empty:
            print(f"没有找到科目编码为 {code_str} 的数据")
            continue
            
        # 应用筛选规则，添加增强的错误处理
        try:
            # 首先检查金额列的数据类型
            if sub_df[debit_column].dtype not in ['int64', 'float64'] or sub_df[credit_column].dtype not in ['int64', 'float64']:
                print("警告：金额列不是数值类型，尝试再次转换...")
                # 再次尝试转换
                try:
                    if sub_df[debit_column].dtype == 'object':
                        sub_df[debit_column] = pd.to_numeric(sub_df[debit_column].astype(str).str.replace(',', '').str.replace('¥', '').str.replace(' ', ''), errors='coerce').fillna(0)
                    if sub_df[credit_column].dtype == 'object':
                        sub_df[credit_column] = pd.to_numeric(sub_df[credit_column].astype(str).str.replace(',', '').str.replace('¥', '').str.replace(' ', ''), errors='coerce').fillna(0)
                    # 转换为float类型
                    sub_df[debit_column] = sub_df[debit_column].astype(float)
                    sub_df[credit_column] = sub_df[credit_column].astype(float)
                except:
                    print("再次转换失败，创建临时排序键...")
                    # 创建临时排序键
                    sub_df['_temp_sort_key'] = range(len(sub_df))
            
            # 现在尝试筛选数据
            if sub_df[debit_credit_column].isnull().all():
                # E列没有内容，筛选L、M列各5个最大的
                try:
                    top_debit = sub_df.nlargest(5, debit_column)
                    top_credit = sub_df.nlargest(5, credit_column)
                    result = pd.concat([top_debit, top_credit]).drop_duplicates()
                    print(f"E列没有内容，筛选L、M列各5个最大值")
                except Exception as e:
                    print(f"筛选L、M列出错: {str(e)}，使用临时排序键")
                    # 使用临时排序键作为备选
                    if '_temp_sort_key' in sub_df:
                        result = sub_df.nlargest(10, '_temp_sort_key')
                    else:
                        result = sub_df.head(10)
            elif (sub_df[debit_credit_column] == '借').any():
                # E列内容是'借'，筛选L列最大的10个
                try:
                    result = sub_df.nlargest(10, debit_column)
                    print(f"E列内容是'借'，筛选L列最大的10个值")
                except Exception as e:
                    print(f"筛选L列出错: {str(e)}，使用临时排序键")
                    if '_temp_sort_key' in sub_df:
                        result = sub_df.nlargest(10, '_temp_sort_key')
                    else:
                        result = sub_df.head(10)
            elif (sub_df[debit_credit_column] == '贷').any():
                # E列内容是'贷'，筛选M列最大的10个
                try:
                    result = sub_df.nlargest(10, credit_column)
                    print(f"E列内容是'贷'，筛选M列最大的10个值")
                except Exception as e:
                    print(f"筛选M列出错: {str(e)}，使用临时排序键")
                    if '_temp_sort_key' in sub_df:
                        result = sub_df.nlargest(10, '_temp_sort_key')
                    else:
                        result = sub_df.head(10)
            else:
                # 默认情况，筛选L列最大的10个
                try:
                    result = sub_df.nlargest(10, debit_column)
                    print(f"E列内容不是'借'或'贷'，默认筛选L列最大的10个值")
                except Exception as e:
                    print(f"默认筛选出错: {str(e)}，使用临时排序键")
                    if '_temp_sort_key' in sub_df:
                        result = sub_df.nlargest(10, '_temp_sort_key')
                    else:
                        result = sub_df.head(10)
            
            # 如果结果为空，使用原始数据前10行
            if result.empty:
                result = sub_df.head(10)
        except Exception as e:
            print(f"筛选数据时出错: {str(e)}")
            # 如果出错，使用原始数据
            result = sub_df.head(10)  # 只取前10行作为默认结果
        
        # 清理临时列
        if '_temp_sort_key' in result:
            result = result.drop('_temp_sort_key', axis=1)
        
        # 创建新的工作表
        sheet_name = get_valid_sheet_name(code_str, workbook)
        
        # 创建或覆盖工作表
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        
        # 创建新工作表
        new_sheet = workbook.create_sheet(sheet_name)
        
        # 首先，如果有原始表格副本，完整复制其格式到新工作表
        if original_sheet:
            try:
                # 复制整个原始表格的格式到新工作表
                # 先创建足够的行和列
                if original_sheet.max_row > new_sheet.max_row:
                    new_sheet.insert_rows(1, original_sheet.max_row - new_sheet.max_row)
                
                # 复制单元格格式和值
                for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row, 
                                                   min_col=1, max_col=original_sheet.max_column):
                    for cell in row:
                        # 确保目标单元格存在
                        target_cell = new_sheet.cell(row=cell.row, column=cell.column)
                        
                        # 复制单元格值（优先复制前2行的模板内容）
                        if cell.row <= 2 and cell.value is not None:
                            target_cell.value = cell.value
                        # 复制标题部分（3-4行）
                        elif cell.row > 2 and cell.row <= 4 and cell.value is not None:
                            target_cell.value = cell.value
                        
                        # 复制单元格格式（对所有行）
                        if cell.has_style:
                            target_cell.font = copy.copy(cell.font)
                            target_cell.border = copy.copy(cell.border)
                            target_cell.fill = copy.copy(cell.fill)
                            target_cell.number_format = cell.number_format
                            target_cell.alignment = copy.copy(cell.alignment)
                            try:
                                target_cell.protection = copy.copy(cell.protection)
                            except:
                                pass
                
                # 复制列宽设置
                for col in original_sheet.column_dimensions:
                    try:
                        if col in new_sheet.column_dimensions:
                            new_sheet.column_dimensions[col].width = original_sheet.column_dimensions[col].width
                        else:
                            new_dim = copy.copy(original_sheet.column_dimensions[col])
                            new_dim.parent = new_sheet
                            new_sheet.column_dimensions[col] = new_dim
                    except:
                        pass
                
                # 复制行高设置
                for row in original_sheet.row_dimensions:
                    try:
                        if row in new_sheet.row_dimensions:
                            new_sheet.row_dimensions[row].height = original_sheet.row_dimensions[row].height
                        else:
                            new_dim = copy.copy(original_sheet.row_dimensions[row])
                            new_dim.parent = new_sheet
                            new_sheet.row_dimensions[row] = new_dim
                    except:
                        pass
                
                # 复制合并单元格
                for merged_range in original_sheet.merged_cells.ranges:
                    try:
                        new_sheet.merge_cells(str(merged_range))
                    except:
                        pass
                
                print(f"成功复制原始表格格式到工作表 '{sheet_name}'")
            except Exception as copy_e:
                print(f"复制格式到新工作表时出错: {str(copy_e)}")
        else:
            # 如果没有找到原始表格副本，使用原有的表头设置
            headers = ['日期', '凭证编号', '业务内容', '明细科目', '对方科目', '金额', '', '核对内容', '', '', '', '', '', '', '', '', '', '', '', '备注']
            sub_headers = ['', '', '', '', '', '借方', '贷方', '1', '2', '3', '4', '5', '6', '7', '8', '', '', '', '', '', '']
            for i, header in enumerate(headers, 1):
                new_sheet.cell(row=3, column=i).value = header
            for i, sub_header in enumerate(sub_headers, 1):
                new_sheet.cell(row=4, column=i).value = sub_header
        
        # 然后调用原有的copy_original_table_to_sheet函数作为备份
        copy_original_table_to_sheet(None, new_sheet, workbook, is_inner_copy=True)
        
        # 填充数据到新工作表（只填充数据，保留格式）
        fill_data(new_sheet, result, df_journal)


def get_valid_sheet_name(name, workbook, max_length=31):
    """获取有效的工作表名称"""
    # 移除非法字符
    invalid_chars = ['\\', '/', ':', '*', '?', '[', ']', '\x00']
    valid_name = ''.join(c for c in name if c not in invalid_chars)
    
    # 限制长度
    valid_name = valid_name[:max_length]
    
    # 如果名称为空，使用默认名称
    if not valid_name:
        valid_name = "Sheet"
    
    # 确保名称不重复
    counter = 1
    temp_name = valid_name
    while temp_name in workbook.sheetnames:
        temp_name = f"{valid_name}_{counter}"
        counter += 1
    
    return temp_name

def safe_set_cell_value(cell, value, fmt=None, is_date_column=False):
    """安全地设置单元格值，保留原有格式"""
    # 导入datetime模块，确保在所有地方都能访问
    from datetime import datetime, timedelta
    # 检查单元格是否是合并单元格的一部分
    # 如果是合并单元格，我们需要获取主单元格（合并区域的左上角单元格）
    if hasattr(cell, 'merged_cell') and cell.merged_cell:
        # 获取工作表
        sheet = cell.parent
        # 遍历所有合并区域
        for merged_range in sheet.merged_cells.ranges:
            # 检查当前单元格是否在这个合并区域内
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                # 找到主单元格（合并区域的左上角单元格）
                cell = sheet.cell(row=min_row, column=min_col)
                break
    # 保存当前单元格格式
    try:
        original_font = copy.copy(cell.font) if cell.font else None
        original_border = copy.copy(cell.border) if cell.border else None
        original_fill = copy.copy(cell.fill) if cell.fill else None
        original_number_format = cell.number_format
        original_alignment = copy.copy(cell.alignment) if cell.alignment else None
        original_protection = copy.copy(cell.protection) if hasattr(cell, 'protection') else None
    except Exception as e:
        print(f"保存单元格格式时出错: {str(e)}")
        # 使用默认值继续
        original_font = None
        original_border = None
        original_fill = None
        original_number_format = 'General'
        original_alignment = None
        original_protection = None
    
    # 尝试设置新值
    try:
        # 处理None或NaN值
        if value is None or (hasattr(value, 'isna') and value.isna()):
            cell.value = ''
        
        # 特殊处理数值类型（优先处理，避免被日期逻辑错误捕获）
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            # 只有在明确是日期列的情况下才尝试日期转换，且增加更严格的限制条件
            if is_date_column:
                try:
                    from datetime import datetime, timedelta
                    # Excel日期基准是1900-01-01
                    # 增加更严格的数值范围检查，避免将金额等大数值错误转换为日期
                    if 1 <= value <= 45000:  # 大约覆盖1900-2100年的日期范围
                        if value >= 1 and value <= 60:  # 处理1900年闰年问题
                            date_value = datetime(1900, 1, 1) + timedelta(days=value - 2)
                        else:
                            date_value = datetime(1900, 1, 1) + timedelta(days=value - 1)
                        # 检查日期是否合理（1900年至今）
                        if 1900 <= date_value.year <= 2100:
                            # 直接转换为字符串格式，确保只显示年月日
                            cell.value = date_value.strftime('%Y-%m-%d')
                            # 设置为文本格式，避免Excel自动转换
                            cell.number_format = '@'
                        else:
                            # 不是有效日期，作为数值处理
                            cell.value = value
                            cell.number_format = fmt if fmt else original_number_format
                    else:
                        # 超出合理日期范围，作为数值处理
                        cell.value = value
                        cell.number_format = fmt if fmt else original_number_format
                except:
                    # 转换失败，作为数值处理
                    cell.value = value
                    cell.number_format = fmt if fmt else original_number_format
            else:
                # 非日期列，直接作为数值处理
                cell.value = value
                cell.number_format = fmt if fmt else original_number_format
        
        # 处理字符串类型
        elif isinstance(value, str):
            # 只有在明确是日期列的情况下才尝试日期转换，且增加更严格的判断
            if is_date_column and len(value.strip()) >= 6:
                try:
                    # 尝试多种日期格式，但避免将凭证编号等格式错误识别为日期
                    # 主要针对年月日格式，忽略可能是凭证编号的格式
                    date_formats = [
                        '%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y',
                        '%Y年%m月%d日', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                        '%d-%m-%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S',
                        '%Y.%m.%d', '%d.%m.%Y', '%m-%d-%Y', '%m/%d/%Y',
                        '%Y年%m月', '%Y-%m', '%Y/%m',
                        # 添加更宽松的解析，但只在日期列中使用
                        '%Y%m%d', '%y%m%d'
                    ]
                    
                    # 增加对凭证编号的特殊判断
                    # 如果是类似"1900-01-01"这样的格式，但可能是凭证编号而不是日期
                    is_possible_certificate = False
                    if len(value) == 10 and value[4] == '-' and value[7] == '-':
                        # 检查是否可能是凭证编号（如以1900开头但不是合理的日期）
                        try:
                            year_part = int(value[:4])
                            if year_part < 2000 and year_part >= 1900:  # 1900-1999年可能是凭证编号
                                # 进一步检查月份和日期部分是否可能为凭证编号的组成部分
                                # 这里只是简单判断，实际应用中可能需要更复杂的逻辑
                                is_possible_certificate = True
                        except:
                            pass
                    
                    if not is_possible_certificate:
                        date_obj = None
                        for date_format in date_formats:
                            try:
                                date_obj = datetime.strptime(value, date_format)
                                break
                            except ValueError:
                                continue
                        
                        if date_obj:
                            # 直接转换为字符串格式，确保只显示年月日
                            # 即使原始字符串包含时间，也只保留日期部分
                            cell.value = date_obj.date().strftime('%Y-%m-%d')
                            # 设置为文本格式，避免Excel自动转换
                            cell.number_format = '@'
                        else:
                            # 不是日期字符串，直接设置值
                            cell.value = value
                    else:
                        # 可能是凭证编号，直接设置值
                        cell.value = value
                except:
                    cell.value = value
            else:
                # 非日期列或太短的字符串，直接设置值
                cell.value = value
        
        # 处理pandas Timestamp类型
        elif isinstance(value, pd.Timestamp):
            # 直接转换为字符串格式，确保只显示年月日
            if pd.notna(value):
                # 确保转换为日期对象并格式化为YYYY-MM-DD
                cell.value = value.date().strftime('%Y-%m-%d')
            else:
                cell.value = ''
            # 设置为文本格式，避免Excel自动转换
            cell.number_format = '@'
        
        # 处理Python datetime类型
        elif isinstance(value, datetime):
            # 直接转换为字符串格式，确保只显示年月日
            if value is not None:
                # 确保转换为日期对象并格式化为YYYY-MM-DD
                cell.value = value.date().strftime('%Y-%m-%d')
            else:
                cell.value = ''
            # 设置为文本格式，避免Excel自动转换
            cell.number_format = '@'
        
        # 其他情况直接设置值
        else:
            try:
                # 尝试直接设置值
                cell.value = value
            except:
                # 如果失败，转换为字符串
                cell.value = str(value)
    except Exception as e:
        print(f"设置单元格值时出错: {str(e)}")
        cell.value = str(value)  # 作为最后的备选方案
    
    # 恢复原有格式 - 增强版处理StyleProxy对象
    try:
        if original_font:
            cell.font = copy.copy(original_font)
    except Exception as e:
        print(f"恢复字体格式时出错: {str(e)}")
    
    try:
        if original_border:
            cell.border = copy.copy(original_border)
    except Exception as e:
        print(f"恢复边框格式时出错: {str(e)}")
    
    try:
        if original_fill:
            cell.fill = copy.copy(original_fill)
    except Exception as e:
        print(f"恢复填充格式时出错: {str(e)}")
    
    # 数字格式是字符串，不需要复制
    try:
        cell.number_format = original_number_format
    except Exception as e:
        print(f"恢复数字格式时出错: {str(e)}")
    
    try:
        if original_alignment:
            cell.alignment = copy.copy(original_alignment)
    except Exception as e:
        print(f"恢复对齐格式时出错: {str(e)}")
    
    try:
        if original_protection:
            cell.protection = copy.copy(original_protection)
    except Exception as e:
        print(f"恢复保护格式时出错: {str(e)}")


def add_table_borders(sheet, min_row, max_row, min_col, max_col):
    """为表格区域添加框线"""
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 为区域内的每个单元格添加边框
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            try:
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
            except Exception as e:
                print(f"添加边框时出错 (行 {row}, 列 {col}): {str(e)}")


def fill_data(sheet, data, df_journal):
    """将数据填充到工作表中，保留原始格式"""
    # 从第5行开始填写数据
    start_row = 5
    max_data_row = start_row  # 跟踪最大数据行
    
    # 按日期倒序排序数据（如果有日期列）
    if hasattr(data, 'sort_values') and not data.empty:
        try:
            # 尝试按日期列排序（假设是第一列）
            # 创建临时日期列用于排序
            data['_date_temp'] = pd.to_datetime(data.iloc[:, 0], errors='coerce')
            
            # 创建辅助排序键，处理无法正确转换为日期的情况
            data['_sort_key'] = range(len(data))  # 默认排序键
            
            # 找出日期列（通常是第一列）
            date_col = data.columns[0]
            
            # 尝试对日期列进行特殊处理，确保能正确识别各种格式的日期
            def parse_date_safely(date_str):
                if pd.isna(date_str):
                    return pd.NaT
                try:
                    # 尝试标准日期解析
                    parsed_date = pd.to_datetime(date_str, errors='coerce')
                    if pd.notna(parsed_date):
                        return parsed_date
                except:
                    pass
                
                try:
                    # 尝试字符串处理，特别处理如"2333-12-18"这样的日期
                    date_str = str(date_str).strip()
                    if len(date_str) >= 8:
                        # 尝试提取年月日信息
                        import re
                        # 匹配 YYYY-MM-DD、YYYY/MM/DD、YYYYMMDD 等格式
                        match = re.search(r'(\d{4})[-/.]?(\d{1,2})[-/.]?(\d{1,2})', date_str)
                        if match:
                            year, month, day = map(int, match.groups())
                            # 检查日期是否有效（扩展年份范围支持未来日期）
                            if 1900 <= year <= 9999 and 1 <= month <= 12 and 1 <= day <= 31:
                                # 直接构建日期字符串，确保能处理未来日期
                                date_str_formatted = f'{year}-{month:02d}-{day:02d}'
                                # 使用dateutil.parser来解析，它对未来日期支持更好
                                from dateutil import parser
                                return pd.Timestamp(parser.parse(date_str_formatted))
                except Exception as e:
                    print(f"解析日期 {date_str} 时出错: {str(e)}")
                return pd.NaT
            
            # 对无法正确转换的日期再试一次特殊处理
            for i, row in data.iterrows():
                if pd.isna(data.at[i, '_date_temp']):
                    date_str = row[date_col]
                    data.at[i, '_date_temp'] = parse_date_safely(date_str)
            
            # 按临时日期列升序排序，对于无法转换为日期的行，使用原始顺序
            data = data.sort_values(by=['_date_temp', '_sort_key'], ascending=[True, True])
            
            # 删除临时列
            data = data.drop(['_date_temp', '_sort_key'], axis=1)
            print(f"已按日期升序排序数据")
        except Exception as e:
            print(f"排序数据时出错: {str(e)}")
            # 排序失败时，尝试按字符串排序作为备选
            try:
                data = data.sort_values(by=data.columns[0], ascending=True)
                print(f"已按字符串升序排序数据作为备选")
            except:
                print(f"备选排序也失败")
    
    # 遍历数据行
    for idx, row in enumerate(data.itertuples(index=False)):
        try:
            current_row = start_row + idx
            max_data_row = max(max_data_row, current_row)
            
            # 确保行不超出Excel限制
            if current_row > 1048576:
                print(f"警告: 行 {current_row} 超出Excel最大行数限制，将被忽略")
                continue
            
            # 确保有足够的行（规则2：自动往下插入行）
            if current_row > sheet.max_row:
                # 插入新行
                sheet.insert_rows(current_row)
                
                # 复制上一行的格式到新插入的行
                if current_row > 1:
                    for col_idx in range(1, min(21, sheet.max_column + 1)):  # 复制前20列的格式
                        try:
                            source_cell = sheet.cell(row=current_row - 1, column=col_idx)
                            target_cell = sheet.cell(row=current_row, column=col_idx)
                            if source_cell.has_style:
                                # 使用深度复制避免StyleProxy问题
                                target_cell.font = copy.copy(source_cell.font)
                                target_cell.border = copy.copy(source_cell.border)
                                target_cell.fill = copy.copy(source_cell.fill)
                                target_cell.number_format = source_cell.number_format
                                target_cell.alignment = copy.copy(source_cell.alignment)
                        except Exception as e:
                            print(f"复制行格式时出错 (行 {current_row}, 列 {col_idx}): {str(e)}")
        
        # 填写数据（规则1），同时保留单元格格式
        # A列：日期（序时账A列）- 明确标记为日期列
            if len(df_journal.columns) > 0:
                date_value = row[0]
                # 直接调用增强版的safe_set_cell_value函数处理所有日期类型
                safe_set_cell_value(sheet.cell(row=current_row, column=1), date_value, 'yyyy-mm-dd', is_date_column=True)
        
        # B列：凭证编号（序时账B列）- 明确标记为非日期列
            if len(df_journal.columns) > 1:
                safe_set_cell_value(sheet.cell(row=current_row, column=2), row[1], None, is_date_column=False)
        
        # C列：业务内容（序时账J列，索引为9）- 明确标记为非日期列
            if len(df_journal.columns) > 9:
                safe_set_cell_value(sheet.cell(row=current_row, column=3), row[9], None, is_date_column=False)
        
        # D列：明细科目（序时账G列，索引为6）- 明确标记为非日期列
            if len(df_journal.columns) > 6:
                safe_set_cell_value(sheet.cell(row=current_row, column=4), row[6], None, is_date_column=False)
        
        # E列：对方科目（序时账I列，索引为8）- 明确标记为非日期列
            if len(df_journal.columns) > 8:
                safe_set_cell_value(sheet.cell(row=current_row, column=5), row[8], None, is_date_column=False)
        
        # F列：借方金额（序时账L列，索引为11）- 明确标记为非日期列
            if len(df_journal.columns) > 11:
                safe_set_cell_value(sheet.cell(row=current_row, column=6), row[11], None, is_date_column=False)
        
        # G列：贷方金额（序时账M列，索引为12）- 明确标记为非日期列
            if len(df_journal.columns) > 12:
                safe_set_cell_value(sheet.cell(row=current_row, column=7), row[12], None, is_date_column=False)
        except Exception as e:
            print(f"填充数据行 {idx+1} 时出错: {str(e)}")
    
    # 添加表格框线，并确保包含备注列（O列，即第15列）
    if max_data_row >= start_row:
        # 为表头添加框线（第3-4行）
        add_table_borders(sheet, 3, 4, 1, 16)
        # 为数据区域添加框线
        add_table_borders(sheet, start_row, max_data_row, 1, 16)
        # 为整个表格区域添加外框（从第1行到最后使用的行）
        # 先确定整个表格区域
        if hasattr(sheet, 'max_row') and sheet.max_row > 0:
            total_rows = sheet.max_row
            # 为整个表格区域添加外框，包含到P列
            for row in range(1, total_rows + 1):
                for col in range(1, 17):  # 覆盖到P列（第16列）
                    cell = sheet.cell(row=row, column=col)
                    # 应用外框
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
    
    # 设置日期列（A列）的宽度为20
    try:
        # 确保A列存在
        if hasattr(sheet, 'column_dimensions'):
            sheet.column_dimensions['A'].width = 20
            print(f"已将日期列（A列）宽度设置为20")
        else:
            # 如果不支持column_dimensions，使用另一种方式设置列宽
            # 注意：这种方式在不同的库实现中可能有所不同
            print("当前库版本可能不支持直接设置列宽")
    except Exception as e:
        print(f"设置日期列宽时出错: {str(e)}")

    # 合并单元格并设置文字居中
    try:
        # 1. 完全合并第一行标题，并将"应交税金检查情况表"中的"应交税金"替换为sheet名称
        if sheet.cell(row=1, column=1).value and "检查情况表" in str(sheet.cell(row=1, column=1).value):
            # 获取当前sheet名称
            sheet_name = sheet.title
            # 替换标题中的"应交税金"为sheet名称
            current_title = str(sheet.cell(row=1, column=1).value)
            new_title = current_title.replace("应交税金", sheet_name)
            sheet.cell(row=1, column=1).value = new_title
            # 合并第一行从A列到P列的所有单元格
            sheet.merge_cells('A1:P1')
            # 设置文字居中
            sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            # 设置字体为加粗和18号
            sheet.cell(row=1, column=1).font = Font(bold=True, size=18)
            print(f"已完全合并并居中第一行标题(A1:P1)，将'应交税金'替换为sheet名称'{sheet_name}'，并设置为加粗和18号字体")
            
            # 合并A2单元格并处理其中的两行文字
            if sheet.cell(row=2, column=1).value:
                # 将A2中的"应交税金"替换为sheet名称
                current_a2_value = str(sheet.cell(row=2, column=1).value)
                new_a2_value = current_a2_value.replace("应交税金", sheet_name)
                # 将"******"替换为"AAA"
                new_a2_value = new_a2_value.replace("******", "AAA")
                sheet.cell(row=2, column=1).value = new_a2_value
                
                sheet.merge_cells('A2:P2')  # 合并A2到P2的整个区域
                # 设置A2数据左对齐
                sheet.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # 设置A2行高为35
                sheet.row_dimensions[2].height = 35
                print(f"已合并A2-P2单元格、设置数据左对齐、设置行高为35，并将'应交税金'替换为sheet名称'{sheet_name}'")
            
            # 合并A18-E18单元格
            sheet.merge_cells('A18:E18')
            # 为A18设置自动换行，按数字换行
            sheet.cell(row=18, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            # 设置A18行高为60
            sheet.row_dimensions[18].height = 60
            print("已合并A18-E18单元格、设置自动换行（按数字换行）并设置行高为60")
            
            # 合并F18-P18单元格
            sheet.merge_cells('F18:P18')
            # 为合并后的F18-P18设置自动换行，按数字换行
            sheet.cell(row=18, column=6).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            print("已合并F18-P18单元格并设置自动换行（按数字换行）")
            
            # 确保表格包含备注列（O列）
            # 检查并设置备注列的宽度，确保能完整显示备注内容
            if 'O' in sheet.column_dimensions:
                sheet.column_dimensions['O'].width = 25  # 设置备注列宽度
                print("已设置备注列（O列）宽度为25，确保完整显示备注内容")
            else:
                # 创建新的列维度对象
                from openpyxl.worksheet.dimensions import ColumnDimension
                new_dim = ColumnDimension(sheet)
                new_dim.width = 25
                sheet.column_dimensions['O'] = new_dim
                print("已创建并设置备注列（O列）宽度为25，确保完整显示备注内容")
            
            # 合并A17-P17单元格，文字靠左
            sheet.merge_cells('A17:P17')
            sheet.cell(row=17, column=1).alignment = Alignment(horizontal='left', vertical='top')
            print("已合并A17-P17单元格并设置文字靠左")
            
            # 合并A19-P19单元格，文字靠左
            sheet.merge_cells('A19:P19')
            sheet.cell(row=19, column=1).alignment = Alignment(horizontal='left', vertical='top')
            print("已合并A19-P19单元格并设置文字靠左")
            
            # 合并A22-P22单元格，文字靠左
            sheet.merge_cells('A22:P22')
            sheet.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='top')
            print("已合并A22-P22单元格并设置文字靠左")
        
        # 2. 合并凭证编号列标题（B3）
        sheet.merge_cells('B3:B4')
        sheet.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并并居中凭证编号列标题")
        
        # 合并A3-A4、C3-C4、D3-D4、E3-E4单元格并设置中文居中
        sheet.merge_cells('A3:A4')
        sheet.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并A3-A4单元格并设置中文居中")
        
        sheet.merge_cells('C3:C4')
        sheet.cell(row=3, column=3).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并C3-C4单元格并设置中文居中")
        
        sheet.merge_cells('D3:D4')
        sheet.cell(row=3, column=4).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并D3-D4单元格并设置中文居中")
        
        sheet.merge_cells('E3:E4')
        sheet.cell(row=3, column=5).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并E3-E4单元格并设置中文居中")
        
        # 3. 合并金额列标题（F3:G3）
        sheet.merge_cells('F3:G3')
        sheet.cell(row=3, column=6).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并并居中金额列标题")
        
        # 4. 处理核对内容区域：保留H3的核对内容，并往右合并一个单元格
        # 保留H3的核对内容，合并范围从H3:N3扩大到H3:O3
        if sheet.cell(row=3, column=8).value and "核对内容" in str(sheet.cell(row=3, column=8).value):
            sheet.merge_cells('H3:O3')  # 往右合并一个单元格，包含备注列
            sheet.cell(row=3, column=8).alignment = Alignment(horizontal='center', vertical='center')
            print("已合并并居中H3的核对内容，并往右合并一个单元格")

        # 合并P3-P4单元格并设置中文居中
        sheet.merge_cells('P3:P4')
        sheet.cell(row=3, column=16).alignment = Alignment(horizontal='center', vertical='center')
        print("已合并P3-P4单元格并设置中文居中")
        
        # 检查H9是否有"核对内容"，如果有则删除
        if sheet.cell(row=9, column=8).value and "核对内容" in str(sheet.cell(row=9, column=8).value):
            sheet.cell(row=9, column=8).value = ""  # 清空H9单元格
            print("已删除H9单元格中的核对内容文本")
        
        # 5. 删除下面红色框框的1-8序号和备注字眼
        # 检查并彻底清空可能存在红色框框的区域
        # 1. 处理H9到P10区域（标准位置）
        sheet.cell(row=9, column=8).value = ""  # 直接清空H9单元格
        sheet.cell(row=9, column=16).value = ""  # 清空P9单元格中的备注字眼
        
        # 清空H10到O10区域的所有单元格
        for row_idx in range(10, 11):  # 第10行
            for col_idx in range(8, 16):  # H到O列
                sheet.cell(row=row_idx, column=col_idx).value = ""  # 强制清空单元格
        
        # 2. 处理数据较少时可能出现的其他红色框框位置（比如第7-11行区域）
        # 检查第7行是否有标题行
        if sheet.cell(row=7, column=1).value and "检查情况表" in str(sheet.cell(row=7, column=1).value):
            # 清空第7-11行中可能的红色框框区域
            for row_idx in range(7, 12):  # 第7-11行
                for col_idx in range(1, 17):  # A到P列
                    sheet.cell(row=row_idx, column=col_idx).value = ""  # 强制清空单元格
            print("已清空数据较少时第7-11行的红色框框内容")
        
        print("已彻底清空所有可能的红色框框区域内容")
            
        # 5. 合并备注列标题（O3和O9）
        if sheet.cell(row=3, column=15).value and "备注" in str(sheet.cell(row=3, column=15).value):
            sheet.merge_cells('O3:O4')
            sheet.cell(row=3, column=15).alignment = Alignment(horizontal='center', vertical='center')
            
        if sheet.cell(row=9, column=15).value and "备注" in str(sheet.cell(row=9, column=15).value):
            sheet.merge_cells('O9:O10')
            sheet.cell(row=9, column=15).alignment = Alignment(horizontal='center', vertical='center')
            print("已合并并居中备注列标题")
    except Exception as e:
        print(f"合并单元格和设置文字居中时出错: {str(e)}")

def copy_original_table_to_sheet(input_file, sheet, workbook, is_inner_copy=False):
    """将原始表格数据写入指定工作表，完整保留格式"""
    if is_inner_copy:
        # 找到模板工作表
        original_sheet = None
        for s in workbook.sheetnames:
            if "模板" in s.lower():
                original_sheet = workbook[s]
                print(f"找到模板工作表: {s}")
                break
        
        # 如果没找到模板工作表，再查找原始表格
        if original_sheet is None:
            for s in workbook.sheetnames:
                if "原始表格" in s:
                    original_sheet = workbook[s]
                    print(f"找到原始表格工作表: {s}")
                    break
        
        if original_sheet:
            # 找到数据区域的最后一行
            data_end_row = 4  # 从第5行开始是数据区域
            for row_num in range(5, sheet.max_row + 1):
                if any(cell.value for cell in sheet[row_num] if cell.column <= 16):
                    data_end_row = row_num
            
            print(f"数据区域最后一行: {data_end_row}")
            
            # 在数据区域后留出几行空白
            start_row = data_end_row + 3
            print(f"开始复制原始表格的起始行: {start_row}")
            
            # 预先扩展工作表行数，避免频繁插入行
            total_needed_rows = start_row + original_sheet.max_row - 1
            if total_needed_rows > sheet.max_row:
                rows_to_add = total_needed_rows - sheet.max_row
                sheet.insert_rows(sheet.max_row + 1, rows_to_add)
                print(f"已扩展工作表行数: {rows_to_add} 行")
            
            # 复制原始表格数据和完整格式 - 使用增强的格式复制方法
            for row_idx in range(1, original_sheet.max_row + 1):
                current_row = start_row + row_idx - 1
                for col_idx in range(1, original_sheet.max_column + 1):
                    source_cell = original_sheet.cell(row=row_idx, column=col_idx)
                    target_cell = sheet.cell(row=current_row, column=col_idx)
                    
                    # 复制单元格值
                    if source_cell.value is not None:
                        target_cell.value = source_cell.value
                    
                    # 深度复制单元格格式 - 安全处理StyleProxy对象
                    if source_cell.has_style:
                        try:
                            # 字体设置
                            if hasattr(source_cell.font, 'name'):
                                target_cell.font = copy.copy(source_cell.font)
                        except Exception as e:
                            print(f"复制字体格式时出错 (行 {row_idx}, 列 {col_idx}): {str(e)}")
                            
                        try:
                            # 边框设置
                            if hasattr(source_cell.border, 'left'):
                                target_cell.border = copy.copy(source_cell.border)
                        except Exception as e:
                            print(f"复制边框格式时出错 (行 {row_idx}, 列 {col_idx}): {str(e)}")
                            
                        try:
                            # 填充设置
                            if hasattr(source_cell.fill, 'patternType'):
                                target_cell.fill = copy.copy(source_cell.fill)
                        except Exception as e:
                            print(f"复制填充格式时出错 (行 {row_idx}, 列 {col_idx}): {str(e)}")
                            
                        try:
                            # 数字格式
                            if hasattr(source_cell, 'number_format'):
                                target_cell.number_format = source_cell.number_format
                        except Exception as e:
                            print(f"复制数字格式时出错 (行 {row_idx}, 列 {col_idx}): {str(e)}")
                            
                        try:
                            # 对齐方式
                            if hasattr(source_cell.alignment, 'horizontal'):
                                target_cell.alignment = copy.copy(source_cell.alignment)
                        except Exception as e:
                            print(f"复制对齐格式时出错 (行 {row_idx}, 列 {col_idx}): {str(e)}")
                            
                        try:
                            # 保护设置
                            if hasattr(source_cell, 'protection') and hasattr(source_cell.protection, 'locked'):
                                target_cell.protection = copy.copy(source_cell.protection)
                        except:
                            pass  # 忽略保护设置的复制错误
            
            # 优化复制列宽设置
            for col_letter in original_sheet.column_dimensions:
                try:
                    if col_letter in original_sheet.column_dimensions:
                        col_dim = original_sheet.column_dimensions[col_letter]
                        if col_letter in sheet.column_dimensions:
                            sheet.column_dimensions[col_letter].width = col_dim.width
                        else:
                            # 创建新的列维度对象，避免直接复制StyleProxy
                            from openpyxl.worksheet.dimensions import ColumnDimension
                            new_dim = ColumnDimension(sheet)
                            new_dim.width = col_dim.width
                            # 复制其他属性
                            if hasattr(col_dim, 'hidden'):
                                new_dim.hidden = col_dim.hidden
                            if hasattr(col_dim, 'bestFit'):
                                new_dim.bestFit = col_dim.bestFit
                            if hasattr(col_dim, 'outlineLevel'):
                                new_dim.outlineLevel = col_dim.outlineLevel
                            sheet.column_dimensions[col_letter] = new_dim
                except Exception as e:
                    print(f"复制列宽时出错 ({col_letter}): {str(e)}")
            
            # 优化复制行高设置
            for row_idx in original_sheet.row_dimensions:
                try:
                    if row_idx in original_sheet.row_dimensions:
                        row_dim = original_sheet.row_dimensions[row_idx]
                        target_row_idx = start_row + row_idx - 1
                        if target_row_idx in sheet.row_dimensions:
                            sheet.row_dimensions[target_row_idx].height = row_dim.height
                        else:
                            # 创建新的行维度对象，避免直接复制StyleProxy
                            from openpyxl.worksheet.dimensions import RowDimension
                            new_dim = RowDimension(sheet)
                            new_dim.height = row_dim.height
                            # 复制其他属性
                            if hasattr(row_dim, 'hidden'):
                                new_dim.hidden = row_dim.hidden
                            if hasattr(row_dim, 'outlineLevel'):
                                new_dim.outlineLevel = row_dim.outlineLevel
                            sheet.row_dimensions[target_row_idx] = new_dim
                except Exception as e:
                    print(f"复制行高时出错 (行 {row_idx}): {str(e)}")
            
            # 改进的合并单元格处理逻辑
            try:
                # 获取合并单元格范围的列表
                merged_ranges = list(original_sheet.merged_cells.ranges)
                print(f"找到 {len(merged_ranges)} 个合并单元格")
                
                if merged_ranges:
                    # 先清除目标工作表中可能的合并单元格冲突
                    try:
                        for merged_range in list(sheet.merged_cells.ranges):
                            sheet.unmerge_cells(str(merged_range))
                    except Exception as clear_e:
                        print(f"清除现有合并单元格时出错: {str(clear_e)}")
                    
                    # 然后复制原始的合并单元格
                    for idx, merged_range in enumerate(merged_ranges):
                        try:
                            # 计算合并单元格在新位置的坐标
                            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(merged_range))
                            
                            # 计算新的合并单元格位置
                            new_min_row = start_row + min_row - 1
                            new_max_row = start_row + max_row - 1
                            
                            # 确保不超出Excel的行列限制
                            if new_min_row > 0 and new_max_row <= 1048576 and min_col > 0 and max_col <= 16384:
                                # 转换为Excel列字母
                                min_col_letter = openpyxl.utils.get_column_letter(min_col)
                                max_col_letter = openpyxl.utils.get_column_letter(max_col)
                                
                                # 创建新的合并单元格范围
                                new_range = f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}"
                                
                                # 执行合并操作
                                sheet.merge_cells(new_range)
                        except Exception as e:
                            # 记录错误但继续处理其他合并单元格
                            print(f"复制合并单元格时出错 ({idx+1}/{len(merged_ranges)}): {str(e)}")
            except Exception as e:
                print(f"处理合并单元格时发生总体错误: {str(e)}")
        else:
            print("未找到原始表格副本工作表，无法复制格式")

def remove_original_copy_sheet(workbook):
    """删除原始表格副本工作表"""
    for sheet_name in list(workbook.sheetnames):
        if "原始表格副本" in sheet_name:
            try:
                del workbook[sheet_name]
                print(f"已删除工作表: {sheet_name}")
            except Exception as e:
                print(f"删除工作表 '{sheet_name}' 时出错: {str(e)}")


# 修改test_copy_table函数以在保存前删除原始表格副本
# 注意：此函数只会在调用时执行，不会影响原函数定义

def modified_test_copy_table():
    # 调用原始的test_copy_table函数
    test_copy_table()


# 重写test_copy_table函数的保存逻辑
original_test_copy_table = test_copy_table

def test_copy_table():
    # 创建输出文件路径的逻辑
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 打开文件选择对话框
    test_file = filedialog.askopenfilename(
        title="选择要处理的Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    
    if not test_file:
        print("用户取消了文件选择，测试终止")
        return
    
    # 创建输出文件路径
    base_name = os.path.basename(test_file)
    name_without_ext = os.path.splitext(base_name)[0]
    output_dir = os.path.dirname(test_file)
    output_file = os.path.join(output_dir, f"{name_without_ext}_处理结果.xlsx")
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # 检查文件扩展名
        file_ext = os.path.splitext(test_file)[1].lower()
        
        # 根据文件类型选择不同的处理方式
        if file_ext == '.xls':
            # 处理.xls文件，需要使用xlrd库
            import xlrd
            
            # 创建新的工作簿作为输出
            wb_output = openpyxl.Workbook()
            wb_output.remove(wb_output.active)  # 删除默认工作表
            
            # 使用xlrd读取.xls文件
            book = xlrd.open_workbook(test_file)
            
            # 复制源文件的所有工作表到输出文件
            for sheet_name in book.sheet_names():
                # 创建新工作表
                ws = wb_output.create_sheet(sheet_name)
                
                # 获取xlrd工作表
                xlrd_sheet = book.sheet_by_name(sheet_name)
                
                # 复制数据
                for row_idx in range(xlrd_sheet.nrows):
                    for col_idx in range(xlrd_sheet.ncols):
                        # 获取单元格值
                        cell_value = xlrd_sheet.cell_value(row_idx, col_idx)
                        
                        # 处理日期类型
                        if xlrd_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            try:
                                # 转换为datetime对象
                                cell_value = datetime(*xlrd.xldate_as_tuple(cell_value, book.datemode))
                            except:
                                # 如果转换失败，保留原始值
                                pass
                        
                        # 写入到openpyxl工作表
                        ws.cell(row=row_idx+1, column=col_idx+1).value = cell_value
            
            # 读取序时账数据（不包含第一行）
            if "序时账" in book.sheet_names():
                xlrd_journal = book.sheet_by_name("序时账")
                # 创建DataFrame
                data = []
                for row_idx in range(1, xlrd_journal.nrows):  # 跳过第一行
                    row_data = []
                    for col_idx in range(xlrd_journal.ncols):
                        cell_value = xlrd_journal.cell_value(row_idx, col_idx)
                        # 处理日期类型
                        if xlrd_journal.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                            try:
                                cell_value = datetime(*xlrd.xldate_as_tuple(cell_value, book.datemode))
                            except:
                                pass
                        row_data.append(cell_value)
                    data.append(row_data)
                # 获取表头
                headers = [xlrd_journal.cell_value(0, col_idx) for col_idx in range(xlrd_journal.ncols)]
                df_journal = pd.DataFrame(data, columns=headers)
            else:
                print("文件中未找到'序时账'工作表")
                return
            
            # 调用process_journal_data函数处理序时账数据
            process_journal_data(df_journal, wb_output)
            
            # 删除原始表格副本工作表
            remove_original_copy_sheet(wb_output)
            
            # 保存输出文件
            wb_output.save(output_file)
            print(f"处理完成，结果已保存到: {output_file}")
        
        else:
            # 处理.xlsx文件
            # 加载原始文件
            original_wb = openpyxl.load_workbook(test_file)
            
            # 创建新的工作簿作为输出
            wb_output = openpyxl.Workbook()
            wb_output.remove(wb_output.active)  # 删除默认工作表
            
            # 复制原始工作簿的所有工作表到输出工作簿
            for sheet_name in original_wb.sheetnames:
                # 创建新工作表
                ws = wb_output.create_sheet(sheet_name)
                
                # 获取原始工作表
                original_sheet = original_wb[sheet_name]
                
                # 复制数据和格式
                for row in original_sheet.iter_rows(min_row=1, max_row=original_sheet.max_row,
                                                  min_col=1, max_col=original_sheet.max_column):
                    for cell in row:
                        # 复制单元格值
                        if cell.value is not None:
                            ws.cell(row=cell.row, column=cell.column).value = cell.value
                        
                        # 复制单元格格式
                        if cell.has_style:
                            try:
                                ws.cell(row=cell.row, column=cell.column).font = copy.copy(cell.font)
                                ws.cell(row=cell.row, column=cell.column).border = copy.copy(cell.border)
                                ws.cell(row=cell.row, column=cell.column).fill = copy.copy(cell.fill)
                                ws.cell(row=cell.row, column=cell.column).number_format = cell.number_format
                                ws.cell(row=cell.row, column=cell.column).alignment = copy.copy(cell.alignment)
                            except:
                                pass
                
                # 复制列宽设置
                for col in original_sheet.column_dimensions:
                    try:
                        if col in ws.column_dimensions:
                            ws.column_dimensions[col].width = original_sheet.column_dimensions[col].width
                        else:
                            ws.column_dimensions[col] = copy.copy(original_sheet.column_dimensions[col])
                            ws.column_dimensions[col].parent = ws
                    except:
                        pass
                
                # 复制行高设置
                for row in original_sheet.row_dimensions:
                    try:
                        if row in ws.row_dimensions:
                            ws.row_dimensions[row].height = original_sheet.row_dimensions[row].height
                        else:
                            ws.row_dimensions[row] = copy.copy(original_sheet.row_dimensions[row])
                            ws.row_dimensions[row].parent = ws
                    except:
                        pass
                
                # 复制合并单元格
                for merged_range in original_sheet.merged_cells.ranges:
                    try:
                        ws.merge_cells(str(merged_range))
                    except:
                        pass
            
            # 读取序时账数据（不包含第一行）
            if "序时账" in original_wb.sheetnames:
                journal_sheet = original_wb["序时账"]
                # 创建DataFrame
                data = []
                headers = []
                
                # 获取表头
                if journal_sheet.max_row > 0:
                    for col_idx in range(1, journal_sheet.max_column + 1):
                        headers.append(journal_sheet.cell(row=1, column=col_idx).value)
                
                # 读取数据行（从第二行开始）
                for row_idx in range(2, journal_sheet.max_row + 1):
                    row_data = []
                    for col_idx in range(1, journal_sheet.max_column + 1):
                        cell = journal_sheet.cell(row=row_idx, column=col_idx)
                        row_data.append(cell.value)
                    data.append(row_data)
                
                df_journal = pd.DataFrame(data, columns=headers)
            else:
                print("文件中未找到'序时账'工作表")
                return
            
            # 调用process_journal_data函数处理序时账数据
            process_journal_data(df_journal, wb_output)
            
            # 删除原始表格副本工作表
            remove_original_copy_sheet(wb_output)
            
            # 保存输出文件
            wb_output.save(output_file)
            print(f"处理完成，结果已保存到: {output_file}")
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        messagebox.showerror("错误", f"处理文件时出错: {str(e)}")


if __name__ == "__main__":
    test_copy_table()