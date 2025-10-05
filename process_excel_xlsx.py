import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import traceback

def process_excel_file(file_path):
    print(f"开始处理文件: {file_path}")
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在")
        return None
    
    try:
        # 读取Excel文件
        xl = pd.ExcelFile(file_path)
        print(f"文件包含工作表: {xl.sheet_names}")
        
        # 读取序时账工作表
        print("\n读取序时账工作表...")
        ledger_df = pd.read_excel(file_path, sheet_name='序时账')
        print(f"序时账数据形状: {ledger_df.shape}")
        print("序时账列名:", list(ledger_df.columns))
        
        # 读取余额表1
        print("\n读取余额表1工作表...")
        balance_df = pd.read_excel(file_path, sheet_name='余额表1')
        print(f"余额表1数据形状: {balance_df.shape}")
        print("余额表1列名:", list(balance_df.columns))
        
        # 使用openpyxl读取序时账以获取背景色信息（不使用data_only=True，以便获取样式）
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ledger_ws = wb['序时账']
        
        # 获取A列（从第2行开始，跳过标题行）背景色为黄色的行索引
        yellow_rows = []
        for row_idx in range(2, ledger_ws.max_row + 1):
            try:
                cell = ledger_ws[f'A{row_idx}']
                # 检查单元格背景色是否为黄色
                fill = cell.fill
                # 黄色的各种表示方式
                if hasattr(fill, 'start_color') and hasattr(fill.start_color, 'index'):
                    color_index = str(fill.start_color.index)
                    # 处理不同格式的黄色表示
                    if 'FFFF00' in color_index or 'FFFF' in color_index and '00' in color_index:
                        # 映射回DataFrame的索引（DataFrame从0开始）
                        yellow_rows.append(row_idx - 2)
            except Exception as e:
                print(f"检查行 {row_idx} 背景色时出错: {e}")
        
        print(f"找到 {len(yellow_rows)} 行背景色为黄色的数据")
        
        # 创建结果DataFrame来存储所有筛选后的数据
        all_filtered_data = []
        
        # 确保余额表1的列名正确
        # 查找F列对应的科目名称列（假设F列是第6列）
        f_column = balance_df.columns[5]  # F列
        print(f"使用列 '{f_column}' 作为余额表1的F列（一级科目名称）")
        
        # 查找G列对应的借方条数列（假设G列是第7列）
        g_column = balance_df.columns[6]  # G列
        print(f"使用列 '{g_column}' 作为余额表1的G列（借方条数）")
        
        # 查找H列对应的贷方条数列（假设H列是第8列）
        h_column = balance_df.columns[7]  # H列
        print(f"使用列 '{h_column}' 作为余额表1的H列（贷方条数）")
        
        # 查找序时账中对应的列
        # 查找F列（一级科目名称，假设F列是第6列）
        ledger_f_column = ledger_df.columns[5]  # F列
        print(f"使用列 '{ledger_f_column}' 作为序时账的F列（一级科目名称）")
        
        # 查找L列（借方发生额，假设L列是第12列）
        ledger_l_column = ledger_df.columns[11]  # L列
        print(f"使用列 '{ledger_l_column}' 作为序时账的L列（借方发生额）")
        
        # 查找M列（贷方发生额，假设M列是第13列）
        ledger_m_column = ledger_df.columns[12]  # M列
        print(f"使用列 '{ledger_m_column}' 作为序时账的M列（贷方发生额）")
        
        # 步骤1: 根据余额表1的内容筛选序时账数据
        print("\n=== 开始筛选数据 ===")
        
        # 创建黄色背景行的标记数组
        is_yellow = [False] * len(ledger_df)
        for idx in yellow_rows:
            if 0 <= idx < len(ledger_df):
                is_yellow[idx] = True
        ledger_df['is_yellow'] = is_yellow
        
        # 遍历余额表1的每一行（不包含第一行标题）
        for idx, row in balance_df.iterrows():
            # 获取一级科目名称
            first_level_name = row[f_column]
            if pd.isna(first_level_name):
                continue
                
            print(f"\n处理一级科目: {first_level_name}")
            
            # 筛选条件1: 该一级科目下的所有数据
            level_data = ledger_df[ledger_df[ledger_f_column] == first_level_name].copy()
            print(f"条件1筛选后: {len(level_data)} 条记录")
            
            # 获取需要筛选的数量
            debit_count = int(row[g_column]) if pd.notna(row[g_column]) else 0
            credit_count = int(row[h_column]) if pd.notna(row[h_column]) else 0
            
            print(f"需要筛选的借方数量: {debit_count}, 贷方数量: {credit_count}")
            
            # 筛选结果存储
            filtered_data = pd.DataFrame()
            
            # 筛选条件2: 按G列数量筛选序时账L列（借方发生额）
            if debit_count > 0:
                # 只筛选有借方发生额的记录
                has_debit = level_data[level_data[ledger_l_column] > 0].copy()
                actual_debit_count = min(debit_count, len(has_debit))
                if actual_debit_count > 0:
                    # 按借方发生额降序排列，取前actual_debit_count条
                    debit_filtered = has_debit.nlargest(actual_debit_count, ledger_l_column)
                    print(f"条件2筛选后，实际筛选借方记录: {len(debit_filtered)} 条")
                    filtered_data = pd.concat([filtered_data, debit_filtered])
                else:
                    print(f"条件2筛选后，没有符合条件的借方记录")
            else:
                print(f"条件2筛选后，debit_count为0，不筛选借方记录")
            
            # 筛选条件3: 按H列数量筛选序时账M列（贷方发生额）
            if credit_count > 0:
                # 只筛选有贷方发生额的记录
                has_credit = level_data[level_data[ledger_m_column] > 0].copy()
                actual_credit_count = min(credit_count, len(has_credit))
                if actual_credit_count > 0:
                    # 按贷方发生额降序排列，取前actual_credit_count条
                    credit_filtered = has_credit.nlargest(actual_credit_count, ledger_m_column)
                    print(f"条件3筛选后，实际筛选贷方记录: {len(credit_filtered)} 条")
                    filtered_data = pd.concat([filtered_data, credit_filtered])
                else:
                    print(f"条件3筛选后，没有符合条件的贷方记录")
            else:
                print(f"条件3筛选后，credit_count为0，不筛选贷方记录")
            
            # 筛选条件4: 添加黄色背景的行
            yellow_data = level_data[level_data['is_yellow'] == True].copy()
            if not yellow_data.empty:
                print(f"条件4筛选后，找到黄色背景记录: {len(yellow_data)} 条")
                # 将黄色背景记录添加到筛选结果中
                filtered_data = pd.concat([filtered_data, yellow_data])
            
            # 去重
            filtered_data = filtered_data.drop_duplicates()
            print(f"该科目最终筛选结果: {len(filtered_data)} 条记录")
            
            # 添加到总结果中
            if not filtered_data.empty:
                all_filtered_data.append(filtered_data)
        
        # 移除临时列
        if 'is_yellow' in ledger_df.columns:
            ledger_df = ledger_df.drop('is_yellow', axis=1)
        
        # 合并所有筛选结果
        if all_filtered_data:
            final_filtered = pd.concat(all_filtered_data)
            print(f"\n总筛选结果: {len(final_filtered)} 条记录")
        else:
            print("\n警告: 没有找到符合条件的记录")
            return None
        
        # 步骤2: 按照一级科目名称拆分到不同工作表
        print("\n=== 开始拆分数据 ===")
        
        # 创建新的Excel文件用于保存结果
        # 定义结果文件路径（使用新的文件名）
        result_file = f"{os.path.splitext(file_path)[0]}_处理结果_new_v6.xlsx"
        
        # 查找模板工作表
        template_loaded = False
        template_data = []
        
        try:
            # 重新加载工作簿以查找模板
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                if '模板' in sheet_name:
                    template_sheet = wb[sheet_name]
                    print(f"找到模板工作表: {template_sheet.title}")
                    
                    # 保存模板数据（只保存单元格值，不保存复杂样式以避免错误）
                    for row in template_sheet.iter_rows():
                        row_values = []
                        for cell in row:
                            row_values.append(cell.value)
                        template_data.append(row_values)
                    
                    template_loaded = True
                    print(f"成功加载模板工作表: {template_sheet.title}")
                    break
            wb.close()
        except Exception as e:
            print(f"加载模板时出错: {e}")
        
        # 创建新工作簿
        result_wb = openpyxl.Workbook()
        # 删除默认的sheet
        default_sheet = result_wb.active
        result_wb.remove(default_sheet)
        
        # 复制筛选后的总数据到一个新工作表
        result_wb.create_sheet('筛选结果')
        result_sheet = result_wb['筛选结果']
        
        # 写入标题行
        for col_idx, col_name in enumerate(final_filtered.columns, 1):
            result_sheet.cell(row=1, column=col_idx, value=col_name)
        
        # 写入数据行
        for row_idx, (_, row) in enumerate(final_filtered.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                result_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 从余额表1的F列获取科目名称的排序顺序
        # 查找F列的列名（可能是'科目名称'或其他名称）
        科目名称列 = None
        for col in balance_df.columns:
            # 检查列内容是否包含科目名称
            if any(balance_df[col].astype(str).str.contains('银行存款|非限定净资产|其他收入|管理费用', na=False)):
                科目名称列 = col
                break
        
        # 如果找到科目名称列，则提取排序顺序
        科目排序顺序 = []
        if 科目名称列:
            # 提取非空的科目名称值
            科目排序顺序 = [str(name).strip() for name in balance_df[科目名称列].dropna() if str(name).strip()]
            print(f"从余额表1提取的科目排序顺序: {科目排序顺序}")
        else:
            print("未找到余额表1中的科目名称列，使用默认排序")
        
        # 按一级科目名称分组
        grouped = final_filtered.groupby(ledger_f_column)
        print(f"共分为 {len(grouped)} 个不同的一级科目")
        
        # 创建按余额表1 F列顺序排序的分组列表
        sorted_groups = []
        # 先添加在余额表1中存在的科目
        for 科目名称 in 科目排序顺序:
            if 科目名称 in grouped.groups:
                sorted_groups.append((科目名称, grouped.get_group(科目名称)))
        # 然后添加余额表1中不存在但在数据中存在的科目
        for name, group_data in grouped:
            if name not in 科目排序顺序:
                sorted_groups.append((name, group_data))
        
        # 按排序后的顺序创建工作表
        for name, group_data in sorted_groups:
            sheet_name = f"{str(name)[:20]}"  # 限制工作表名长度
            
            # 创建新工作表
            new_sheet = result_wb.create_sheet(sheet_name)
            
            # 如果加载了模板，复制模板内容
            if template_loaded and template_data:
                # 填充模板内容（值）
                for row_idx, row_values in enumerate(template_data, 1):
                    for col_idx, value in enumerate(row_values, 1):
                        if col_idx <= len(row_values):
                            new_sheet.cell(row=row_idx, column=col_idx, value=value)
                print(f"已应用模板格式到工作表: {sheet_name}")
            
            print(f"\n创建工作表: {sheet_name}, 记录数: {len(group_data)}")
            
            # 对数据先按日期（A列）排序，再按凭证编号（B列）排序
            # 将数据转换为可排序的格式
            sorted_data = []
            for _, row in group_data.iterrows():
                # 获取日期和凭证编号用于排序
                date_value = row.iloc[0] if len(ledger_df.columns) > 0 else ''
                voucher_value = row.iloc[1] if len(ledger_df.columns) > 1 else ''
                sorted_data.append((date_value, voucher_value, row))
            
            # 排序：先按日期，再按凭证编号
            sorted_data.sort(key=lambda x: (x[0], x[1]))
            
            # 从第5行开始填写已排序的数据
            for data_idx, (_, _, row) in enumerate(sorted_data, 5):
                try:
                    # A5：日期（序时账A列，第1列）- 只保留年月日
                    if len(ledger_df.columns) > 0:
                        date_value = row.iloc[0]
                        if pd.notna(date_value):
                            # 尝试将日期格式化为只包含年月日
                            try:
                                if isinstance(date_value, str):
                                    # 如果是字符串，尝试解析为日期
                                    date_obj = pd.to_datetime(date_value)
                                    new_sheet[f'A{data_idx}'] = date_obj.strftime('%Y-%m-%d')
                                else:
                                    # 如果是日期类型，直接格式化
                                    new_sheet[f'A{data_idx}'] = date_value.strftime('%Y-%m-%d')
                            except:
                                # 如果格式化失败，保留原始值
                                new_sheet[f'A{data_idx}'] = str(date_value).split()[0]  # 尝试简单分割去除时分秒
                        else:
                            new_sheet[f'A{data_idx}'] = ''
                    
                    # B5：凭证编号（序时账B列，第2列）
                    if len(ledger_df.columns) > 1:
                        new_sheet[f'B{data_idx}'] = row.iloc[1] if pd.notna(row.iloc[1]) else ''
                    
                    # C5：业务内容（序时账J列，第10列）
                    if len(ledger_df.columns) > 9:
                        new_sheet[f'C{data_idx}'] = row.iloc[9] if pd.notna(row.iloc[9]) else ''
                    
                    # D5：明细科目（序时账G列，第7列）
                    if len(ledger_df.columns) > 6:
                        new_sheet[f'D{data_idx}'] = row.iloc[6] if pd.notna(row.iloc[6]) else ''
                    
                    # E5：对方科目（序时账I列，第9列）
                    if len(ledger_df.columns) > 8:
                        new_sheet[f'E{data_idx}'] = row.iloc[8] if pd.notna(row.iloc[8]) else ''
                    
                    # F5：金额（根据借贷方填写）
                    debit_amount = row.iloc[11] if (len(ledger_df.columns) > 11 and pd.notna(row.iloc[11])) else 0
                    credit_amount = row.iloc[12] if (len(ledger_df.columns) > 12 and pd.notna(row.iloc[12])) else 0
                    new_sheet[f'F{data_idx}'] = debit_amount if debit_amount > 0 else credit_amount
                    
                    # G5：借方（序时账L列，第12列）
                    if len(ledger_df.columns) > 11:
                        new_sheet[f'G{data_idx}'] = debit_amount
                    
                    # H5：留空（之前是贷方）
                    new_sheet[f'H{data_idx}'] = ''
                    
                    # I-O：留空，等待手动填写
                    for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O']:
                        new_sheet[f'{col}{data_idx}'] = ''
                    
                    # P5：核对内容（留空，等待手动填写）
                    new_sheet[f'P{data_idx}'] = ''
                    
                    # Q5：备注（序时账N列，第14列）
                    if len(ledger_df.columns) > 13:
                        new_sheet[f'Q{data_idx}'] = row.iloc[13] if pd.notna(row.iloc[13]) else ''
                except Exception as e:
                    print(f"填充数据时出错: {e}")
                    traceback.print_exc()
        
        # 保存结果文件
        result_wb.save(result_file)
        
        print(f"\n处理完成! 结果已保存至: {result_file}")
        return result_file
        
    except Exception as e:
        print(f"处理文件时出错: {e}")
        traceback.print_exc()
        return None

if __name__ == "__main__":
    # 处理目标文件
    target_file = r"c:\Users\54293\Desktop\6、南通狼山大圣爱心慈善基金会.xlsx"
    process_excel_file(target_file)
    print("\n处理完成，请查看生成的文件。")